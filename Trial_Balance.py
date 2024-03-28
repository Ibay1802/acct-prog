import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
import pandas as pd
import openpyxl as ox
from openpyxl import load_workbook
from tkinter import *
from tkcalendar import Calendar
import os
import locale
import datetime
import numpy as np
from num2words import num2words
import math
from pandastable import Table
import sys

#Trial Balance
def tb ():
    window_tb = tk.Tk()
    window_tb.title("Trial Balance")
    window_tb.geometry("700x400+400+10")
    window_tb.configure(bg='#333333')
    frame_tb=tk.Frame(window_tb,bg='#333333')
    frame_tb.pack(fill=BOTH,expand=True)

    def start_date(event):
        global cal, date_window
        date_window = Toplevel()
        date_window.grab_set()
        date_window.title('Choose Date')
        date_window.geometry('250x220+400+70')
        cal = Calendar(date_window, selectmode="day", date_pattern="dd/mm/yyyy")
        cal.place(x=0, y=0)
        submit_btn = Button(date_window, text="submit", command=grabstart_date)
        submit_btn.place(x=80, y=190)
    def grabstart_date():
        startdate_entry.delete(0,END)
        startdate_entry.insert(0, cal.get_date())
        date_window.destroy()

    def end_date(event):
        global cal, date_window
        date_window = Toplevel()
        date_window.grab_set()
        date_window.title('Choose Date')
        date_window.geometry('250x220+400+70')
        cal = Calendar(date_window, selectmode="day", date_pattern="dd/mm/yyyy")
        cal.place(x=0, y=0)
        submit_btn = Button(date_window, text="submit", command=grab_enddate)
        submit_btn.place(x=80, y=190)
    def grab_enddate():
        enddate_entry.delete(0, END)
        enddate_entry.insert(0, cal.get_date())
        date_window.destroy()

    df_tb = ox.load_workbook("lap_keu.xlsx")
    ws_tb = df_tb['Neraca Saldo']
    row_index = 1
    def generate_tb ():
        try:
            #date = startdate_entry.get()
            #year = date[-4:]
            #file = f'database_{year}.xlsx'
            file="database.xlsx"
            df_data = pd.read_excel(file)
            global row_index 
            start_date = pd.to_datetime(startdate_entry.get(),format="%d/%m/%Y", errors="coerce").date()
            end_date = pd.to_datetime(enddate_entry.get(),format="%d/%m/%Y", errors="coerce").date()
            range_date = pd.date_range(start = start_date, end= end_date)
            data_tanggal = df_data.loc[(df_data[df_data.columns[0]].isin (range_date))]
            grup_akun = data_tanggal.groupby(data_tanggal.columns[4])[["Debet", "Kredit"]].sum()
            saldo_akun = grup_akun.sort_values(by=[grup_akun.columns[0]])
            #print(grup_akun)
            #print(saldo_akun)
        # Iterate over each row of ws_tb and write the values of saldo_akun in the specified columns
        
            for row in ws_tb.iter_rows(min_row=1, min_col=4, max_col=5):
                for cell in row:
                    cell.value = None
            row_index = 1
            for row in ws_tb.iter_rows(min_row=1, values_only=True):
                no_akun = row[0]
                if no_akun in saldo_akun.index:
                    ws_tb.cell(row=row_index, column=4, value=saldo_akun.loc[no_akun][0])
                    ws_tb.cell(row=row_index, column=5, value=saldo_akun.loc[no_akun][1])
                row_index += 1
            df_tb.save("lap_keu.xlsx")
            startdate_entry.delete(0,END)
            enddate_entry.delete(0,END)
            messagebox.showinfo(title="Generate success", message="Generate Success")
            window_tb.destroy()
        except:
            messagebox.showerror("Error",message="Generate Failed")

    #Design Trial Balance
    tb_label = tk.Label(frame_tb,text="Trial Balance",bg='#333333',fg='#FFFFFF',font=["arial",20],justify="center")
    tb_label.place(x=200,y=20)

    #Start Date
    startdate_label = tk.Label(frame_tb, text="Dari Tanggal",bg ='#333333', fg='white',font=["arial",14])
    startdate_label.place(x=20, y=100)
    startdate_entry = tk.Entry(frame_tb, highlightthickness=0,bg="white",fg="black",font=("arial",14),width=10)
    startdate_entry.place(x=150, y=100)
    startdate_entry.insert(0, "dd/mm/yyyy")
    startdate_entry.bind("<1>",start_date)
    #End date
    enddate_label = tk.Label(frame_tb, text="- Sampai Tanggal",bg ='#333333', fg='white',font=["arial",14])
    enddate_label.place(x=300, y=100)
    enddate_entry = tk.Entry(frame_tb, highlightthickness=0,bg="white",fg="black",font=("arial",14),width=10)
    enddate_entry.place(x=480, y=100)
    enddate_entry.insert(0, "dd/mm/yyyy")
    enddate_entry.bind("<1>",end_date)

    #Button
    generate_button = tk.Button(frame_tb,text="Generate",bg="White",fg="black",font=["arial",12],width=45,command=generate_tb)
    generate_button.place(x=20,y=150)

    window_tb.mainloop()