#laba Rugi
import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
import pandas as pd
import openpyxl as ox
from openpyxl import load_workbook
from tkcalendar import Calendar
from tkinter import *
import os
import locale
import datetime
import numpy as np
from num2words import num2words
import math
from pandastable import Table
import sys
from tkinter import scrolledtext
import io


def laba_rugi ():
    window_pl = tk.Tk()
    window_pl.title("Laba Rugi")
    window_pl.geometry("1080x800+400+5")
    window_pl.configure(bg='#333333')
    frame_pl=tk.Frame(window_pl,bg="white")
    frame_pl.pack(fill=BOTH,expand=True)

    def startdate(event):
        global cal, date_window
        date_window = Toplevel()
        date_window.grab_set()
        date_window.title('Choose Date')
        date_window.geometry('250x220+800+70')
        cal = Calendar(date_window, selectmode="day", date_pattern="dd/mm/yyyy")
        cal.place(x=0, y=0)
        submit_btn = Button(date_window, text="submit", command=grabstart_date)
        submit_btn.place(x=80, y=190)
    def grabstart_date():
        start_entry.delete(0,END)
        start_entry.insert(0, cal.get_date())
        date_window.destroy()

    def enddate(event):
        global cal, date_window
        date_window = Toplevel()
        date_window.grab_set()
        date_window.title('Choose Date')
        date_window.geometry('250x220+800+70')
        cal = Calendar(date_window, selectmode="day", date_pattern="dd/mm/yyyy")
        cal.place(x=0, y=0)
        submit_btn = Button(date_window, text="submit", command=grab_enddate)
        submit_btn.place(x=80, y=190)
    def grab_enddate():
        end_entry.delete(0, END)
        end_entry.insert(0, cal.get_date())
        date_window.destroy()

    df_tb = ox.load_workbook("lap_keu.xlsx")
    ws_tb = df_tb['Neraca Saldo']
    row_index = 1
    def generate_tb ():
        try:
            #date = start_entry.get()
            #year = date[-4:]
            #file = f'database_{year}.xlsx'
            file = "database.xlsx"
            df_data = pd.read_excel(file)
            global row_index 
            start_date = pd.to_datetime(start_entry.get(),format="%d/%m/%Y", errors="coerce").date()
            end_date = pd.to_datetime(end_entry.get(),format="%d/%m/%Y", errors="coerce").date()
            range_date = pd.date_range(start = start_date, end= end_date)
            data_tanggal = df_data.loc[(df_data[df_data.columns[0]].isin (range_date))]
            grup_akun = data_tanggal.groupby(data_tanggal.columns[4])[["Debet", "Kredit"]].sum()
            saldo_akun = grup_akun.sort_values(by=[grup_akun.columns[0]])
            #print(grup_akun)
            #print(saldo_akun)
        # Iterate over each row of ws_tb and write the values of saldo_akun in the specified columns       
            for row in ws_tb.iter_rows(min_row=1, min_col=4, max_col=5):
                for cell in row:
                    cell.value = None
            row_index = 1
            for row in ws_tb.iter_rows(min_row=1, values_only=True):
                no_akun = row[0]
                if no_akun in saldo_akun.index:
                    ws_tb.cell(row=row_index, column=4, value=saldo_akun.loc[no_akun][0])
                    ws_tb.cell(row=row_index, column=5, value=saldo_akun.loc[no_akun][1])
                row_index += 1
            df_tb.save("lap_keu.xlsx")
            messagebox.showinfo(title="Generate success", message="Generate Success")
        except:
            messagebox.showerror("Error",message="Generate Failed")

    def get_rev_sum(ws_tb, min_row, max_row, col_idx):
        debit = sum(int(cell.value) for row in ws_tb.iter_rows(min_row=min_row, max_row=max_row, min_col=4, max_col=4) for cell in row if cell.value is not None)
        credit = sum(int(cell.value) for row in ws_tb.iter_rows(min_row=min_row, max_row=max_row, min_col=5, max_col=5) for cell in row if cell.value is not None)
        return  credit - debit

    def get_cost_sum(ws_tb, min_row, max_row, col_idx):
        debit = sum(int(cell.value) for row in ws_tb.iter_rows(min_row=min_row, max_row=max_row, min_col=4, max_col=4) for cell in row if cell.value is not None)
        credit = sum(int(cell.value) for row in ws_tb.iter_rows(min_row=min_row, max_row=max_row, min_col=5, max_col=5) for cell in row if cell.value is not None)
        return  debit + credit

    #Fungsi Profit & Loss
    def profit_loss ():
        
        proj_rev = get_rev_sum(ws_tb,133,133,4)
        train_rev = get_rev_sum(ws_tb,134,134,4)
        tekser_rev = get_rev_sum(ws_tb,135,135,4)
        other_rev = get_rev_sum(ws_tb,136,136,4)
        total_rev = proj_rev + train_rev + tekser_rev +other_rev
        
        giro_rev = get_rev_sum(ws_tb,138,138,4)
        kurs_rev = get_rev_sum(ws_tb,139,139,4)
        depst_rev = get_rev_sum(ws_tb,140,140,4)
        oth_rev = get_rev_sum(ws_tb,141,141,4)
        total_otherrev = giro_rev + kurs_rev + depst_rev +oth_rev

        proj_exp = get_cost_sum(ws_tb,144,147,4)
        train_exp = get_cost_sum(ws_tb,149,152,4)
        tekser_exp = get_cost_sum(ws_tb,154,157,4)
        total_cos = proj_exp + train_exp + tekser_exp
        gross_profit = total_rev - total_cos
        row = 158
        adum_exp = get_cost_sum(ws_tb, min_row=row, max_row=row, col_idx=4)
        row += 1
        sal_exp = get_cost_sum(ws_tb, min_row=row, max_row=row, col_idx=4)
        row += 1
        pan_exp = get_cost_sum(ws_tb, min_row=row, max_row=row, col_idx=4)
        row += 1
        kon_exp = get_cost_sum(ws_tb, min_row=row, max_row=row, col_idx=4)
        row += 1
        tun_exp = get_cost_sum(ws_tb, min_row=row, max_row=row, col_idx=4)
        row += 1
        uty_exp = get_cost_sum(ws_tb, min_row=row, max_row=row, col_idx=4)
        row += 1
        offic_exp = get_cost_sum(ws_tb, min_row=row, max_row=row, col_idx=4)
        row += 1
        tax21_exp = get_cost_sum(ws_tb, min_row=row, max_row=row, col_idx=4)
        row += 1
        veh_exp = get_cost_sum(ws_tb, min_row=row, max_row=row, col_idx=4)
        row += 1
        stat_exp = get_cost_sum(ws_tb, min_row=row, max_row=row, col_idx=4)
        row += 1
        trans_exp = get_cost_sum(ws_tb, min_row=row, max_row=row, col_idx=4)
        row += 1
        emptrain_exp = get_cost_sum(ws_tb, min_row=row, max_row=row, col_idx=4)
        row += 1
        travel_exp = get_cost_sum(ws_tb, min_row=row, max_row=row, col_idx=4)
        row += 1
        ins_exp = get_cost_sum(ws_tb, min_row=row, max_row=row, col_idx=4)
        row += 1
        bpjs_exp = get_cost_sum(ws_tb, min_row=row, max_row=row, col_idx=4)
        row += 1
        build_depre = get_cost_sum(ws_tb, min_row=row, max_row=row, col_idx=4)
        row += 1
        furni_depre = get_cost_sum(ws_tb, min_row=row, max_row=row, col_idx=4)
        row += 1
        equip_depre = get_cost_sum(ws_tb, min_row=row, max_row=row, col_idx=4)
        row += 1
        vehi_depre = get_cost_sum(ws_tb, min_row=row, max_row=row, col_idx=4)
        row += 1
        total_adm = (adum_exp+ sal_exp+ pan_exp+ kon_exp+ tun_exp+ uty_exp+ offic_exp+ 
                    tax21_exp+ veh_exp+ stat_exp+ trans_exp+ emptrain_exp+ travel_exp+
                    ins_exp+ bpjs_exp+ build_depre + furni_depre + equip_depre + vehi_depre)
        opra_prof = gross_profit - total_adm
        row2=179
        Prov_exp = get_cost_sum(ws_tb, min_row=row, max_row=row, col_idx=4)
        row += 1
        int_exp = get_cost_sum(ws_tb, min_row=row, max_row=row, col_idx=4)
        row += 1
        bank_exp = get_cost_sum(ws_tb, min_row=row, max_row=row, col_idx=4)
        row += 1
        leasint_exp = get_cost_sum(ws_tb, min_row=row, max_row=row, col_idx=4)
        row += 1
        bad_exp = get_cost_sum(ws_tb, min_row=row, max_row=row, col_idx=4)
        row += 1
        kurs_exp = get_cost_sum(ws_tb, min_row=row, max_row=row, col_idx=4)
        row += 1
        oth_exp = get_cost_sum(ws_tb, min_row=row, max_row=row, col_idx=4)
        row += 1
        total_other = Prov_exp+int_exp+bank_exp+leasint_exp+bad_exp+kurs_exp+oth_exp
        inc_beftax = opra_prof + total_otherrev -total_other
        fin_tax = get_rev_sum(ws_tb,188,188,4)
        comp_tax = get_rev_sum(ws_tb,187,187,4)
        other_tax = get_rev_sum(ws_tb,189,191,4)
        tot_tax = fin_tax + comp_tax+ other_tax
        inc_afttax = inc_beftax + tot_tax

        output_text = scrolledtext.ScrolledText(frame_pl, width=80, height=40)
        output_text.place(x = 20, y=100)

        class TextRedirector(io.TextIOBase):
            def __init__(self, widget):
                self.widget = widget
            def write(self, string):
                self.widget.insert(tk.END, string)
                self.widget.see(tk.END)
                
        sys.stdout = TextRedirector(output_text)
        
        locale.setlocale(locale.LC_ALL, "")
        print(f"Pendapatan Usaha")
        print(f"Pendapatan Projek             : {locale.format_string('%d', proj_rev, grouping=True):>12}")
        print(f"Pendapatan Training           : {locale.format_string('%d', train_rev, grouping=True):>12}")
        print(f"Pendapatan Teknikal Service   : {locale.format_string('%d', tekser_rev, grouping=True):>12}")
        print(f"Other Revenue                 : {locale.format_string('%d', other_rev, grouping=True):>12}")
        print(f"                               ______________")
        print(f"Total Revenue                 : {locale.format_string('%d', total_rev, grouping=True):>12}")
        print(f"Biaya Usaha")
        print(f"Biaya Usaha Projek            : {locale.format_string('%d',proj_exp,grouping=True):>12}")
        print(f"Biaya Usaha Training          : {locale.format_string('%d', train_exp, grouping=True):>12}")
        print(f"Biaya Usaha Teknikal Service  : {locale.format_string('%d', tekser_exp, grouping=True):>12}")
        print(f"                               ______________")
        print(f"Total Biaya Usaha             : {locale.format_string('%d', total_cos, grouping=True):>12}")
        print(f"                               ______________")
        print(f"Gross Profit                  : {locale.format_string('%d', gross_profit, grouping=True):>12}")
        print(f"Biaya Administrasi dan Umum")
        print(f"Biaya Umum                    : {locale.format_string('%d',adum_exp , grouping=True):>12}")
        print(f"Biaya Gaji                    : {locale.format_string('%d',sal_exp , grouping=True):>12}")
        print(f"Biaya Keb.Dapur               : {locale.format_string('%d',pan_exp , grouping=True):>12}")
        print(f"Biaya Konsumsi                : {locale.format_string('%d',kon_exp , grouping=True):>12}")
        print(f"Biaya Tunjangan               : {locale.format_string('%d',tun_exp , grouping=True):>12}")
        print(f"Biaya Listrik, Air & Tlp      : {locale.format_string('%d',uty_exp , grouping=True):>12}")
        print(f"Biaya Maint.Kantor            : {locale.format_string('%d',offic_exp , grouping=True):>12}")
        print(f"Biaya PPh 21                  : {locale.format_string('%d',tax21_exp , grouping=True):>12}")
        print(f"Biaya Maint. Kendaraan        : {locale.format_string('%d',veh_exp , grouping=True):>12}")
        print(f"Biaya ATK & Perl.Kantor       : {locale.format_string('%d',stat_exp , grouping=True):>12}")
        print(f"Biaya Trasportasi             : {locale.format_string('%d',trans_exp , grouping=True):>12}")
        print(f"Biaya Training Karyawan       : {locale.format_string('%d',emptrain_exp , grouping=True):>12}")
        print(f"Biaya Perjalan Dinas Adm      : {locale.format_string('%d',travel_exp , grouping=True):>12}")
        print(f"Biaya Asuransi                : {locale.format_string('%d',ins_exp , grouping=True):>12}")
        print(f"Biaya BPJS                    : {locale.format_string('%d',bpjs_exp , grouping=True):>12}")
        print(f"Biaya Penyusutan Bangunan     : {locale.format_string('%d',build_depre , grouping=True):>12}")
        print(f"Biaya Penyusutan Furniture    : {locale.format_string('%d',furni_depre , grouping=True):>12}")
        print(f"Biaya Penyusutan Perl. Kantor : {locale.format_string('%d',equip_depre , grouping=True):>12}")
        print(f"Biaya Penyusutan Kendaraan    : {locale.format_string('%d',vehi_depre , grouping=True):>12}")
        print(f"                               ______________")
        print(f"Total Biaya Adm & Umum        : {locale.format_string('%d', total_adm, grouping=True):>12}")
        print(f"                               ______________")
        print(f"Laba Operasional              : {locale.format_string('%d',opra_prof , grouping=True):>12}")
        print(f"Pendapatan & Biaya Lain-lain")
        print(f"Pendapatan Lain-lain")
        print(f"Jasa Giro Bank                : {locale.format_string('%d',giro_rev , grouping=True):>12}")
        print(f"Laba Selisih Kurs             : {locale.format_string('%d',kurs_rev , grouping=True):>12}")
        print(f"Bunga Deposito Bank           : {locale.format_string('%d',depst_rev , grouping=True):>12}")
        print(f"Pendapatan Lainnya            : {locale.format_string('%d',oth_rev , grouping=True):>12}")
        print(f"                               ______________")
        print(f"Total Pendapatan Lain-lain    : {locale.format_string('%d', total_otherrev, grouping=True):>12}")
        print(f"Biaya Lain-lain")
        print(f"Biaya Provisi Bank            : {locale.format_string('%d',Prov_exp , grouping=True):>12}")
        print(f"Biaya Bunga Bank              : {locale.format_string('%d',int_exp , grouping=True):>12}")
        print(f"Biaya Administrasi Bank       : {locale.format_string('%d',bank_exp , grouping=True):>12}")
        print(f"Biaya Bunga Leasing           : {locale.format_string('%d',leasint_exp , grouping=True):>12}")
        print(f"Biaya Bad Debt                : {locale.format_string('%d',bad_exp , grouping=True):>12}")
        print(f"Biaya Selisih Kurs            : {locale.format_string('%d',kurs_exp , grouping=True):>12}")
        print(f"Biaya Lainnya                 : {locale.format_string('%d',oth_exp , grouping=True):>12}")
        print(f"                               ______________")
        print(f"Total Biaya Lain-lain         : {locale.format_string('%d', total_other, grouping=True):>12}")
        print(f"                               ______________")
        print(f"Laba Sebelum Pajak            : {locale.format_string('%d', inc_beftax, grouping=True):>12}")
        print(f"                               ______________")
        print(f"Biaya Pajak")
        print(f"Biaya Pajak Final             : {locale.format_string('%d', fin_tax, grouping=True):>12}")
        print(f"Biaya Pajak Badan             : {locale.format_string('%d', comp_tax, grouping=True):>12}")
        print(f"Biaya Pajak Lainnya           : {locale.format_string('%d', other_tax, grouping=True):>12}")
        print(f"                               ______________")
        print(f"Total Biaya  Pajak            : {locale.format_string('%d', tot_tax, grouping=True):>12}")
        print(f"                               ______________")
        print(f"Laba Setelah Pajak            : {locale.format_string('%d', inc_afttax, grouping=True):>12}")
        print(f"                               ==============")

    sys.stdout = sys.__stdout__

    template_lr = ox.load_workbook("template_LR.xlsx")
    ws = template_lr["Laba-Rugi"]
    def print_pl ():
        start = start_entry.get()
        end = end_entry.get()
        periode_lr = "Periode "+ start +" - "+ end
        ws.cell(row=3,column=1).value = periode_lr
        proj_rev = get_rev_sum(ws_tb,133,133,4)
        train_rev = get_rev_sum(ws_tb,134,134,4)
        tekser_rev = get_rev_sum(ws_tb,135,135,4)
        other_rev = get_rev_sum(ws_tb,136,136,4)
        total_rev = proj_rev + train_rev + tekser_rev +other_rev
        row_num = 5
        for rev in [proj_rev, train_rev,tekser_rev, other_rev,total_rev] :
            ws.cell(row=row_num,column=3).value = int(rev)
            row_num += 1
        proj_exp = get_cost_sum(ws_tb,144,147,4)
        train_exp = get_cost_sum(ws_tb,149,152,4)
        tekser_exp = get_cost_sum(ws_tb,154,157,4)
        totalcos_exp = proj_exp + train_exp + tekser_exp
        row_num2=12
        for exp in [proj_exp, train_exp,tekser_exp,totalcos_exp] :
            ws.cell(row=row_num2,column=3).value = int(exp)
            row_num2 += 1
        gross_profit = total_rev - totalcos_exp
        ws.cell(row=15,column=3).value = int(gross_profit)
        row = 158
        adum_exp = get_cost_sum(ws_tb, min_row=row, max_row=row, col_idx=4)
        row += 1
        sal_exp = get_cost_sum(ws_tb, min_row=row, max_row=row, col_idx=4)
        row += 1
        pan_exp = get_cost_sum(ws_tb, min_row=row, max_row=row, col_idx=4)
        row += 1
        kon_exp = get_cost_sum(ws_tb, min_row=row, max_row=row, col_idx=4)
        row += 1
        tun_exp = get_cost_sum(ws_tb, min_row=row, max_row=row, col_idx=4)
        row += 1
        uty_exp = get_cost_sum(ws_tb, min_row=row, max_row=row, col_idx=4)
        row += 1
        offic_exp = get_cost_sum(ws_tb, min_row=row, max_row=row, col_idx=4)
        row += 1
        tax21_exp = get_cost_sum(ws_tb, min_row=row, max_row=row, col_idx=4)
        row += 1
        veh_exp = get_cost_sum(ws_tb, min_row=row, max_row=row, col_idx=4)
        row += 1
        stat_exp = get_cost_sum(ws_tb, min_row=row, max_row=row, col_idx=4)
        row += 1
        trans_exp = get_cost_sum(ws_tb, min_row=row, max_row=row, col_idx=4)
        row += 1
        emptrain_exp = get_cost_sum(ws_tb, min_row=row, max_row=row, col_idx=4)
        row += 1
        travel_exp = get_cost_sum(ws_tb, min_row=row, max_row=row, col_idx=4)
        row += 1
        ins_exp = get_cost_sum(ws_tb, min_row=row, max_row=row, col_idx=4)
        row += 1
        bpjs_exp = get_cost_sum(ws_tb, min_row=row, max_row=row, col_idx=4)
        row += 1
        build_exp = get_cost_sum(ws_tb, min_row=row, max_row=row, col_idx=4)
        row += 1
        furni_exp = get_cost_sum(ws_tb, min_row=row, max_row=row, col_idx=4)
        row += 1
        equip_exp = get_cost_sum(ws_tb, min_row=row, max_row=row, col_idx=4)
        row += 1
        vehi_exp = get_cost_sum(ws_tb, min_row=row, max_row=row, col_idx=4)
        row += 1
        totaladm_exp = (adum_exp+ sal_exp+ pan_exp+ kon_exp+ tun_exp+ uty_exp+ offic_exp+ 
                    tax21_exp+ veh_exp+ stat_exp+ trans_exp+ emptrain_exp+ travel_exp+
                    ins_exp+ bpjs_exp+ build_exp + furni_exp + equip_exp + vehi_exp)
        row_num3=19
        for exp in [adum_exp,sal_exp,pan_exp,kon_exp,tun_exp,uty_exp,offic_exp,tax21_exp,veh_exp,stat_exp,trans_exp,
                    emptrain_exp,travel_exp,ins_exp,bpjs_exp,build_exp,furni_exp,equip_exp,vehi_exp,totaladm_exp] :
            ws.cell(row=row_num3,column=3).value = int(exp)
            row_num3 += 1
        opra_prof = gross_profit - totaladm_exp
        ws.cell(row=40,column=3).value = int(opra_prof)
        giro_rev = get_rev_sum(ws_tb,138,138,4)
        kurs_rev = get_rev_sum(ws_tb,139,139,4)
        depst_rev = get_rev_sum(ws_tb,140,140,4)
        oth_rev = get_rev_sum(ws_tb,141,141,4)
        totalother_rev = giro_rev + kurs_rev + depst_rev +oth_rev
        row_num4=44
        for rev in [giro_rev,kurs_rev,depst_rev,oth_rev,totalother_rev] :
            ws.cell(row=row_num4,column=3).value = int(rev)
            row_num4 += 1

        Prov_exp = get_cost_sum(ws_tb, min_row=row, max_row=row, col_idx=4)
        row += 1
        int_exp = get_cost_sum(ws_tb, min_row=row, max_row=row, col_idx=4)
        row += 1
        bank_exp = get_cost_sum(ws_tb, min_row=row, max_row=row, col_idx=4)
        row += 1
        leasint_exp = get_cost_sum(ws_tb, min_row=row, max_row=row, col_idx=4)
        row += 1
        bad_exp = get_cost_sum(ws_tb, min_row=row, max_row=row, col_idx=4)
        row += 1
        kurs_exp = get_cost_sum(ws_tb, min_row=row, max_row=row, col_idx=4)
        row += 1
        oth_exp = get_cost_sum(ws_tb, min_row=row, max_row=row, col_idx=4)
        row += 1
        totalother_exp = Prov_exp+int_exp+bank_exp+leasint_exp+bad_exp+kurs_exp+oth_exp
        row_num5=50
        for exp in [Prov_exp,int_exp,bank_exp,leasint_exp,bad_exp,kurs_exp,oth_exp,totalother_exp] :
            ws.cell(row=row_num5,column=3).value = int(exp)
            row_num5 += 1
        inc_beftax = opra_prof + totalother_rev - totalother_exp
        ws.cell(row=40,column=3).value = int(inc_beftax)
        fin_tax = get_rev_sum(ws_tb,188,188,4)
        comp_tax = get_rev_sum(ws_tb,187,187,4)
        other_tax = get_rev_sum(ws_tb,189,191,4)
        tot_tax = fin_tax + comp_tax + other_tax
        row_num6=62
        for tax in [fin_tax,comp_tax,other_tax,tot_tax] :
            ws.cell(row=row_num6,column=3).value = int(tax)
            row_num6 += 1
        inc_afttax = inc_beftax + tot_tax
        ws.cell(row=67,column=3).value = int(inc_afttax)

        template_lr.save('template_LR.xlsx')
        os.startfile('template_LR.xlsx')
        messagebox.showinfo(title="Export Laba Rugi", message="Export Lap. Laba-Rugi Success")

    #Design Profit & Loss
    pl_label = tk.Label(frame_pl,text="Laporan Laba Rugi",bg="white",fg='black',font=["arial",20],justify="center")
    pl_label.place(x=200,y=10)

    #Periode
    per_label=tk.Label(frame_pl,text="Periode",bg="white",fg="black",font=("arial",12))
    per_label.place(x=135,y=60)
    per_label=tk.Label(frame_pl,text=" - ",bg="white",fg="black",font=("arial",12))
    per_label.place(x=298,y=60)
    start_entry = tk.Entry(frame_pl, highlightthickness=0,bg="white",fg="black",font=("arial",12),width=10)
    start_entry.place(x=200, y=60)
    start_entry.insert(0, "dari tanggal")
    start_entry.bind("<1>",startdate)
    end_entry = tk.Entry(frame_pl, highlightthickness=0,bg="white",fg="black",font=("arial",12),width=15)
    end_entry.place(x=320, y=60)
    end_entry.insert(0, "sampai tanggal")
    end_entry.bind("<1>",enddate)

    #Button
    generate_button = tk.Button(frame_pl,text="Generate",bg="black",fg="white",font=["arial",12],width=25,command=generate_tb)
    generate_button.place(x=800,y=50)
    pl_button = tk.Button(frame_pl,text="Tampilkan Laba-Rugi",bg="black",fg="white",font=["arial",12],width=25,command=profit_loss)
    pl_button.place(x=800,y=90)
    print_button = tk.Button(frame_pl,text="Print Laba-Rugi",bg="black",fg="white",font=["arial",12],width=25,command=print_pl)
    print_button.place(x=800,y=130)

    window_pl.mainloop()