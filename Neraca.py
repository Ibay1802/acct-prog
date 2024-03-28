import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
import pandas as pd
import openpyxl as ox
from openpyxl import load_workbook
from tkcalendar import Calendar
from tkinter import *
import os
import locale
import datetime
import numpy as np
from num2words import num2words
import math
from pandastable import Table
import sys

#Balance Sheet
def balance_sheet () :
    window_bs = tk.Tk()
    window_bs.title("Neraca")
    window_bs.geometry("1080x800+400+5")
    window_bs.configure(bg='#333333')
    frame_bs=tk.Frame(window_bs,bg='#333333')
    frame_bs.pack(fill=BOTH,expand=True)

    def as_date(event):
        global cal, date_window
        date_window = Toplevel()
        date_window.grab_set()
        date_window.title('Choose Date')
        date_window.geometry('250x220+400+70')
        cal = Calendar(date_window, selectmode="day", date_pattern="dd/mm/yyyy")
        cal.place(x=0, y=0)
        submit_btn = Button(date_window, text="submit", command=grabas_date)
        submit_btn.place(x=80, y=190)
    def grabas_date():
        date_entry.delete(0,END)
        date_entry.insert(0, cal.get_date())
        date_window.destroy()

    df_tb = ox.load_workbook("lap_keu.xlsx")
    ws_tb = df_tb['Neraca Saldo']
    row_index = 1
    def generate_tb ():
        #try:
            #date = date_entry.get()
            #year = date[-4:]
            #file = f'database_{year}.xlsx'
            file="database.xlsx"
            df_data = pd.read_excel(file)
            global row_index 
            as_date = pd.to_datetime(date_entry.get(),format="%d/%m/%Y", errors="coerce").date()
            as_datetime = pd.to_datetime(as_date)
            date_data = pd.to_datetime(df_data[Tanggal])
            data_tanggal = df_data.loc[(date_data <= (as_datetime))]
            #data_tanggal = df_data.loc[(df_data[df_data.columns[0]] <= (as_datetime))]
            grup_akun = data_tanggal.groupby(data_tanggal.columns[4])[["Debet", "Kredit"]].sum()
            saldo_akun = grup_akun.sort_values(by=[grup_akun.columns[0]])
        # Iterate over each row of ws_tb and write the values of saldo_akun in the specified columns
            row_index = 1
            for row in ws_tb.iter_rows(min_row=1, values_only=True):
                no_akun = row[0]
                if no_akun in saldo_akun.index:
                    ws_tb.cell(row=row_index, column=4, value=saldo_akun.loc[no_akun][0])
                    ws_tb.cell(row=row_index, column=5, value=saldo_akun.loc[no_akun][1])
                row_index += 1
            df_tb.save("lap_keu.xlsx")
            #messagebox.showinfo(title="Generate success", message="Generate Success")
        #except:
            #messagebox.showerror("Error",message="Generate Failed")
         #Fungsi Neraca
    def balance_sheet ():
        def get_asset_sum(ws_tb, min_row, max_row, col_idx):
            awal = sum(int(cell.value) for row in ws_tb.iter_rows(min_row=min_row, max_row=max_row, min_col=3, max_col=3) for cell in row if cell.value is not None)
            debit = sum(int(cell.value) for row in ws_tb.iter_rows(min_row=min_row, max_row=max_row, min_col=4, max_col=4) for cell in row if cell.value is not None)
            credit = sum(int(cell.value) for row in ws_tb.iter_rows(min_row=min_row, max_row=max_row, min_col=5, max_col=5) for cell in row if cell.value is not None)
            return awal + debit - credit
        #Cash
        cash = get_asset_sum(ws_tb, 4, 15, 3)
        cash_entry.delete(0,END)
        cash_entry.insert(0,cash)
        #AR
        ar = get_asset_sum(ws_tb, 17, 20, 3)
        ar_entry.delete(0,END)
        ar_entry.insert(0,ar)
        #AR Other
        otherar = get_asset_sum(ws_tb, 22, 24, 3)
        otherar_entry.delete(0,END)
        otherar_entry.insert(0,otherar)
        #Uang Muka
        uangmuka = get_asset_sum(ws_tb, 26, 30, 3)
        uangmuka_entry.delete(0,END)
        uangmuka_entry.insert(0,uangmuka)
        #Advance Cost
        bdm = get_asset_sum(ws_tb, 39, 47, 3)
        bdm_entry.delete(0,END)
        bdm_entry.insert(0,bdm)
        #Advance Tax
        pdm = get_asset_sum(ws_tb, 32, 37, 3)
        pdm_entry.delete(0,END)
        pdm_entry.insert(0,pdm)
        #Total current Asset
        tca= cash + ar + otherar + uangmuka +bdm + pdm
        totalca_entry.delete(0,END)
        totalca_entry.insert(0,tca)

        #Fixed Asset
        #Equipment
        equipment = get_asset_sum(ws_tb, 49, 49, 3)
        equipment_entry.delete(0,END)
        equipment_entry.insert(0,equipment)
        #furniture
        furniture = get_asset_sum(ws_tb, 50, 50, 3)
        furniture_entry.delete(0,END)
        furniture_entry.insert(0,furniture)
        #vehicle
        vehicle = get_asset_sum(ws_tb, 51, 51, 3)
        vehicle_entry.delete(0,END)
        vehicle_entry.insert(0,vehicle)
        #Total FA
        tfa= equipment + furniture + vehicle
        totalfa_entry.delete(0,END)
        totalfa_entry.insert(0,tfa)
        #Accum. Depre. FA
        #accm.equipment
        accmequip = get_asset_sum(ws_tb, 54, 54, 3)
        accmequip_entry.delete(0,END)
        accmequip_entry.insert(0,accmequip)
        #accm.Furniture
        accmfurniture = get_asset_sum(ws_tb, 55, 55, 3)
        accmfurniture_entry.delete(0,END)
        accmfurniture_entry.insert(0,accmfurniture)
        #accm.Vehicle
        accmvehicle = get_asset_sum(ws_tb, 56, 56, 3)
        accmvehicle_entry.delete(0,END)
        accmvehicle_entry.insert(0,accmvehicle)
        #Total ACDFA
        acdfa= accmequip + accmfurniture + accmvehicle
        totalaccm_entry.delete(0,END)
        totalaccm_entry.insert(0,acdfa)
        #Nilai Buku
        nb= tfa + acdfa
        bookvalue_entry.delete(0,END)
        bookvalue_entry.insert(0,nb)
        #Total Asset
        ta= tca + nb
        totalasset_entry.delete(0,END)
        totalasset_entry.insert(0,ta)
        ##LIABILITIES
        #Current Liabilities
        def get_Passiva_sum(ws_tb, min_row, max_row, col_idx):
            awal = sum(int(cell.value) for row in ws_tb.iter_rows(min_row=min_row, max_row=max_row, min_col=3, max_col=3) for cell in row if cell.value is not None)
            debit = sum(int(cell.value) for row in ws_tb.iter_rows(min_row=min_row, max_row=max_row, min_col=4, max_col=4) for cell in row if cell.value is not None)
            credit = sum(int(cell.value) for row in ws_tb.iter_rows(min_row=min_row, max_row=max_row, min_col=5, max_col=5) for cell in row if cell.value is not None)
            return awal - debit + credit
        #Account Payable
        ap = get_Passiva_sum(ws_tb, 60, 62, 3)
        ap_entry.delete(0,END)
        ap_entry.insert(0,ap)
        #Prepaid Cost
        bymd = get_Passiva_sum(ws_tb, 64, 81, 3)
        bymd_entry.delete(0,END)
        bymd_entry.insert(0,bymd)
        #Tax Payable
        taxpay = get_Passiva_sum(ws_tb, 83, 91, 3)
        taxpay_entry.delete(0,END)
        taxpay_entry.insert(0,taxpay)
        #Bank Loan
        bankpay = get_Passiva_sum(ws_tb, 93, 96, 3)
        bankpay_entry.delete(0,END)
        bankpay_entry.insert(0,bankpay)
        #Advance Income
        advance = get_Passiva_sum(ws_tb, 98, 107, 3)
        advance_entry.delete(0,END)
        advance_entry.insert(0,advance)
        #Affiliated & Shareholder Loan
        affiliate = get_Passiva_sum(ws_tb, 109, 112, 3)
        affiliate_entry.delete(0,END)
        affiliate_entry.insert(0,affiliate)
        #Total Current Liabilities
        totalcl= ap + bymd + taxpay + bankpay + advance + affiliate
        totalcl_entry.delete(0,END)
        totalcl_entry.insert(0,totalcl)

        #Long Term Liabilities Bank
        longbank = get_Passiva_sum(ws_tb, 114, 116, 3)
        longbank_entry.delete(0,END)
        longbank_entry.insert(0,longbank)
        #Long Term Liabilities Affiliation & Share Holder
        longowner = get_Passiva_sum(ws_tb, 118, 121, 3)
        longowner_entry.delete(0,END)
        longowner_entry.insert(0,longowner)
        #Total Long Term Liabilities
        totallong= longbank +longowner
        totallong_entry.delete(0,END)
        totallong_entry.insert(0,totallong)
        #Total Liabilities
        totalpay= totalcl + totallong
        totalpay_entry.delete(0,END)
        totalpay_entry.insert(0,totalpay)
        #EQUITY
        #Share Holde Equity
        ownequity =get_Passiva_sum(ws_tb, 123, 123, 3)
        ownequity_entry.delete(0,END)
        ownequity_entry.insert(0,ownequity)
        #Retained Earning
        re =get_Passiva_sum(ws_tb, 125, 126, 3)
        re_entry.delete(0,END)
        re_entry.insert(0,re)
        #Current PL
        currpl_awal = sum(int(cell.value) for row in ws_tb.iter_rows(min_row=128, max_row=128, min_col=3, max_col=3) for cell in row if cell.value is not None)
        currpl_debit = sum(int(cell.value) for row in ws_tb.iter_rows(min_row=133, max_row=191, min_col=4, max_col=4) for cell in row if cell.value is not None)
        currpl_credit = sum(int(cell.value) for row in ws_tb.iter_rows(min_row=133, max_row=191, min_col=5, max_col=5) for cell in row if cell.value is not None)
        currpl =currpl_awal - currpl_debit + currpl_credit
        currpl_entry.delete(0,END)
        currpl_entry.insert(0,currpl)
        #Total Equity
        totaleq= ownequity + re + currpl
        totaleq_entry.delete(0,END)
        totaleq_entry.insert(0,totaleq)
        #Total Equity + Liabilities
        totalpasiv= totaleq + totalpay
        totalpasiv_entry.delete(0,END)
        totalpasiv_entry.insert(0,totalpasiv)
        if ta-totalpasiv == 0 :
                totalasset_entry.config(background='green')
                totalpasiv_entry.config(background='green')
        else :
                totalasset_entry.config(background='red')
                totalpasiv_entry.config(background='red')

    template_neraca = ox.load_workbook('Neraca Template.xlsx')
    sheet = template_neraca['NERACA']
    def export_excel ():
        per_neraca = "Per Tanggal" + date_entry.get()
        sheet.cell(row=3,column=1).value = per_neraca
        row_num = 6
        for entry in [cash_entry, ar_entry, otherar_entry, uangmuka_entry, bdm_entry, pdm_entry]:
            sheet.cell(row=row_num, column=2).value = entry.get()
            row_num += 1 
        row_num2= 15
        for entry in [equipment_entry, furniture_entry, vehicle_entry]:
            sheet.cell(row=row_num2, column=2).value = entry.get()
            row_num2 += 1 
        row_num3= 20
        for entry in [accmequip_entry, accmfurniture_entry, accmvehicle_entry]:
            sheet.cell(row=row_num3, column=2).value = entry.get()
            row_num3 += 1 
        row_num= 6
        for entry in [ap_entry, bymd_entry, taxpay_entry,bankpay_entry,advance_entry,affiliate_entry]:
            sheet.cell(row=row_num, column=5).value = entry.get()
            row_num += 1 
        row_num2= 15
        for entry in [longbank_entry, longowner_entry]:
            sheet.cell(row=row_num2, column=5).value = entry.get()
            row_num2 += 1 
        row_num4= 22
        for entry in [ownequity_entry, re_entry,currpl_entry]:
            sheet.cell(row=row_num4, column=5).value = entry.get()
            row_num4 += 1 

        #new_file = "Neraca" + " " + per_neraca+" "+datetime.datetime.now().strftime("%Y-%m-%d")+".xlsx"
        template_neraca.save('Neraca Template.xlsx')
        os.startfile('Neraca Template.xlsx')
        window_bs.destroy()

    neraca_label = tk.Label(frame_bs, text= "NERACA",bg='#333333',fg='#FFFFFF',font=["arial",20],justify="center")
    neraca_label.place(x=230,y=10)
    date_label = tk.Label(frame_bs, text="Per ",bg ='#333333', fg='white',font=["arial",12],justify="center")
    date_label.place(x=220, y=50)
    date_entry = tk.Entry(frame_bs, highlightthickness=0,bg='#333333',fg='white',font=("arial",12),width=10)
    date_entry.place(x=250, y=50)
    date_entry.insert(0, "dd/mm/yyyy")
    date_entry.bind("<1>",as_date)

    aktiva_label = tk.Label(frame_bs, text= "AKTIVA",bg='#333333',fg='#FFFFFF',font=["arial",12])
    aktiva_label.place(x=20,y=120)
    aktiva_lancar = tk.Label(frame_bs, text= "Aktiva Lancar :",bg='#333333',fg='#FFFFFF',font=["arial",12])
    aktiva_lancar.place(x=20,y=150)
    cash_label = tk.Label(frame_bs, text= "Kas Dan Setara Kas",bg='#333333',fg='#FFFFFF',font=["arial",12])
    cash_label.place(x=50,y=180)
    cash_entry = tk.Entry(frame_bs,bg='#333333',fg='white',font=("arial",12),justify="right")
    cash_entry.place(x=230,y=180)

    ar_label = tk.Label(frame_bs, text= "Piutang Usaha",bg='#333333',fg='#FFFFFF',font=["arial",12])
    ar_label.place(x=50,y=210)
    ar_entry = tk.Entry(frame_bs,bg='#333333',fg='white',font=("arial",12),justify="right")
    ar_entry.place(x=230,y=210)

    otherar_label = tk.Label(frame_bs, text= "Piutang Lainnya",bg='#333333',fg='#FFFFFF',font=["arial",12])
    otherar_label.place(x=50,y=240)
    otherar_entry = tk.Entry(frame_bs,bg='#333333',fg='white',font=("arial",12),justify="right")
    otherar_entry.place(x=230,y=240)

    uangmuka_label = tk.Label(frame_bs, text= "Uang Muka",bg='#333333',fg='#FFFFFF',font=["arial",12])
    uangmuka_label.place(x=50,y=270)
    uangmuka_entry = tk.Entry(frame_bs,bg='#333333',fg='white',font=("arial",12),justify="right")
    uangmuka_entry.place(x=230,y=270)

    bdm_label = tk.Label(frame_bs, text= "Biaya dibayar dimuka",bg='#333333',fg='#FFFFFF',font=["arial",12])
    bdm_label.place(x=50,y=300)
    bdm_entry = tk.Entry(frame_bs,bg='#333333',fg='white',font=("arial",12),justify="right")
    bdm_entry.place(x=230,y=300)

    pdm_label = tk.Label(frame_bs, text= "Pajak dibayar dimuka",bg='#333333',fg='#FFFFFF',font=["arial",12])
    pdm_label.place(x=50,y=330)
    pdm_entry = tk.Entry(frame_bs,bg='#333333',fg='white',font=("arial",12),justify="right")
    pdm_entry.place(x=230,y=330)

    totalca_label = tk.Label(frame_bs, text= "Total Aktiva Lancar",bg='#333333',fg='#FFFFFF',font=["arial",12])
    totalca_label.place(x=20,y=360)
    totalca_entry = tk.Entry(frame_bs,bg='grey',fg='white',font=("arial",12),justify="right")
    totalca_entry.place(x=230,y=360)

    aktiva_tetap = tk.Label(frame_bs, text= "Aktiva Tetap:",bg='#333333',fg='#FFFFFF',font=["arial",12])
    aktiva_tetap.place(x=20,y=390)
    equipment_label = tk.Label(frame_bs, text= "Peralatan Kantor",bg='#333333',fg='#FFFFFF',font=["arial",12])
    equipment_label.place(x=50,y=420)
    equipment_entry = tk.Entry(frame_bs,bg='#333333',fg='white',font=("arial",12),justify="right")
    equipment_entry.place(x=230,y=420)

    furniture_label = tk.Label(frame_bs, text= "Furniture Kantor",bg='#333333',fg='#FFFFFF',font=["arial",12])
    furniture_label.place(x=50,y=450)
    furniture_entry = tk.Entry(frame_bs,bg='#333333',fg='white',font=("arial",12),justify="right")
    furniture_entry.place(x=230,y=450)

    vehicle_label = tk.Label(frame_bs, text= "Kendaraan",bg='#333333',fg='#FFFFFF',font=["arial",12])
    vehicle_label.place(x=50,y=480)
    vehicle_entry = tk.Entry(frame_bs,bg='#333333',fg='white',font=("arial",12),justify="right")
    vehicle_entry.place(x=230,y=480)

    totalfa_label = tk.Label(frame_bs, text= "Total Aktiva Tetap",bg='#333333',fg='#FFFFFF',font=["arial",12])
    totalfa_label.place(x=20,y=510)
    totalfa_entry = tk.Entry(frame_bs,bg='grey',fg='white',font=("arial",12),justify="right")
    totalfa_entry.place(x=230,y=510)

    accm_depre = tk.Label(frame_bs, text= "Akumulasi Penyusutan:",bg='#333333',fg='#FFFFFF',font=["arial",12])
    accm_depre.place(x=20,y=540)

    accmequip_label = tk.Label(frame_bs, text= "Akum. Peny. Peralatan",bg='#333333',fg='#FFFFFF',font=["arial",12])
    accmequip_label.place(x=50,y=570)
    accmequip_entry = tk.Entry(frame_bs,bg='#333333',fg='white',font=("arial",12),justify="right")
    accmequip_entry.place(x=230,y=570)

    accmfurniture_label = tk.Label(frame_bs, text= "Akum. Peny. furniture",bg='#333333',fg='#FFFFFF',font=["arial",12])
    accmfurniture_label.place(x=50,y=600)
    accmfurniture_entry = tk.Entry(frame_bs,bg='#333333',fg='white',font=("arial",12),justify="right")
    accmfurniture_entry.place(x=230,y=600)
    accmvehicle_label = tk.Label(frame_bs, text= "Akum. Peny. Kendaraan",bg='#333333',fg='#FFFFFF',font=["arial",12])
    accmvehicle_label.place(x=50,y=630)
    accmvehicle_entry = tk.Entry(frame_bs,bg='#333333',fg='white',font=("arial",12),justify="right")
    accmvehicle_entry.place(x=230,y=630)

    totalaccm_label = tk.Label(frame_bs, text= "Total akum. Peny.",bg='#333333',fg='#FFFFFF',font=["arial",12])
    totalaccm_label.place(x=20,y=660)
    totalaccm_entry = tk.Entry(frame_bs,bg='grey',fg='white',font=("arial",12),justify="right")
    totalaccm_entry.place(x=230,y=660)

    bookvalue_label = tk.Label(frame_bs, text= "Nilai Buku",bg='#333333',fg='#FFFFFF',font=["arial",12])
    bookvalue_label.place(x=20,y=690)
    bookvalue_entry = tk.Entry(frame_bs,bg='grey',fg='white',font=("arial",12),justify="right")
    bookvalue_entry.place(x=230,y=690)

    totalasset_label = tk.Label(frame_bs, text= "Total Aktiva",bg='#333333',fg='#FFFFFF',font=["arial",14])
    totalasset_label.place(x=20,y=720)
    totalasset_entry = tk.Entry(frame_bs,bg='blue',fg='white',font=("arial",14),justify="right")
    totalasset_entry.place(x=230,y=720)

    payable_label = tk.Label(frame_bs, text= "KEWAJIBAN",bg='#333333',fg='#FFFFFF',font=["arial",12])
    payable_label.place(x=500,y=120)
    current_ap = tk.Label(frame_bs, text= "Hutang Lancar :",bg='#333333',fg='#FFFFFF',font=["arial",12])
    current_ap.place(x=500,y=150)
    ap_label = tk.Label(frame_bs, text= "Hutang Usaha",bg='#333333',fg='#FFFFFF',font=["arial",12])
    ap_label.place(x=530,y=180)
    ap_entry = tk.Entry(frame_bs,bg='#333333',fg='white',font=("arial",12),justify="right")
    ap_entry.place(x=700,y=180)

    bymd_label = tk.Label(frame_bs, text= "Hutang Biaya",bg='#333333',fg='#FFFFFF',font=["arial",12])
    bymd_label.place(x=530,y=210)
    bymd_entry = tk.Entry(frame_bs,bg='#333333',fg='white',font=("arial",12),justify="right")
    bymd_entry.place(x=700,y=210)

    taxpay_label = tk.Label(frame_bs, text= "Hutang Pajak",bg='#333333',fg='#FFFFFF',font=["arial",12])
    taxpay_label.place(x=530,y=240)
    taxpay_entry = tk.Entry(frame_bs,bg='#333333',fg='white',font=("arial",12),justify="right")
    taxpay_entry.place(x=700,y=240)

    bankpay_label = tk.Label(frame_bs, text= "Hutang Bank",bg='#333333',fg='#FFFFFF',font=["arial",12])
    bankpay_label.place(x=530,y=270)
    bankpay_entry = tk.Entry(frame_bs,bg='#333333',fg='white',font=("arial",12),justify="right")
    bankpay_entry.place(x=700,y=270)

    advance_label = tk.Label(frame_bs, text= "Pendapatan Dimuka",bg='#333333',fg='#FFFFFF',font=["arial",12])
    advance_label.place(x=530,y=300)
    advance_entry = tk.Entry(frame_bs,bg='#333333',fg='white',font=("arial",12),justify="right")
    advance_entry.place(x=700,y=300)

    affiliate_label = tk.Label(frame_bs, text= "Hutang Affil & Owner",bg='#333333',fg='#FFFFFF',font=["arial",12])
    affiliate_label.place(x=530,y=330)
    affiliate_entry = tk.Entry(frame_bs,bg='#333333',fg='white',font=("arial",12),justify="right")
    affiliate_entry.place(x=700,y=330)

    totalcl_label = tk.Label(frame_bs, text= "Total Hutang Lancar",bg='#333333',fg='#FFFFFF',font=["arial",12])
    totalcl_label.place(x=500,y=360)
    totalcl_entry = tk.Entry(frame_bs,bg='grey',fg='white',font=("arial",12),justify="right")
    totalcl_entry.place(x=700,y=360)

    long_ap = tk.Label(frame_bs, text= "Hutang Jangka Panjang :",bg='#333333',fg='#FFFFFF',font=["arial",12])
    long_ap.place(x=500,y=390)
    longbank_label = tk.Label(frame_bs, text= "Hutang bank JP",bg='#333333',fg='#FFFFFF',font=["arial",12])
    longbank_label.place(x=530,y=420)
    longbank_entry = tk.Entry(frame_bs,bg='#333333',fg='white',font=("arial",12),justify="right")
    longbank_entry.place(x=700,y=420)

    longowner_label = tk.Label(frame_bs, text= "Hutang Affil & Owner JP ",bg='#333333',fg='#FFFFFF',font=["arial",12])
    longowner_label.place(x=530,y=450)
    longowner_entry = tk.Entry(frame_bs,bg='#333333',fg='white',font=("arial",12),justify="right")
    longowner_entry.place(x=700,y=450)

    totallong_label = tk.Label(frame_bs, text= "Total Hut. Jangka Panjang",bg='#333333',fg='#FFFFFF',font=["arial",12])
    totallong_label.place(x=500,y=480)
    totallong_entry = tk.Entry(frame_bs,bg='grey',fg='white',font=("arial",12),justify="right")
    totallong_entry.place(x=700,y=480)

    totalpay_label = tk.Label(frame_bs, text= "Total Kewajiban",bg='#333333',fg='#FFFFFF',font=["arial",14])
    totalpay_label.place(x=500,y=510)
    totalpay_entry = tk.Entry(frame_bs,bg='blue',fg='white',font=("arial",14),justify="right")
    totalpay_entry.place(x=700,y=510)

    equity_label = tk.Label(frame_bs, text= "MODAL",bg='#333333',fg='#FFFFFF',font=["arial",12])
    equity_label.place(x=500,y=540)
    ownequity_label = tk.Label(frame_bs, text= "Modal Pemilik",bg='#333333',fg='#FFFFFF',font=["arial",12])
    ownequity_label.place(x=530,y=570)
    ownequity_entry = tk.Entry(frame_bs,bg='#333333',fg='white',font=("arial",12),justify="right")
    ownequity_entry.place(x=700,y=570)

    re_label = tk.Label(frame_bs, text= "Laba Ditahan",bg='#333333',fg='#FFFFFF',font=["arial",12])
    re_label.place(x=530,y=600)
    re_entry = tk.Entry(frame_bs,bg='#333333',fg='white',font=("arial",12),justify="right")
    re_entry.place(x=700,y=600)

    currpl_label = tk.Label(frame_bs, text= "Laba Rugi Berjalan",bg='#333333',fg='#FFFFFF',font=["arial",12])
    currpl_label.place(x=530,y=630)
    currpl_entry = tk.Entry(frame_bs,bg='#333333',fg='white',font=("arial",12),justify="right")
    currpl_entry.place(x=700,y=630)

    totaleq_label = tk.Label(frame_bs, text= "Total Modal",bg='#333333',fg='#FFFFFF',font=["arial",14])
    totaleq_label.place(x=500,y=660)
    totaleq_entry = tk.Entry(frame_bs,bg='blue',fg='white',font=("arial",14),justify="right")
    totaleq_entry.place(x=700,y=660)

    totalpasiv_label = tk.Label(frame_bs, text= "Total Kwjb + Modal",bg='#333333',fg='#FFFFFF',font=["arial",14])
    totalpasiv_label.place(x=500,y=720)
    totalpasiv_entry = tk.Entry(frame_bs,bg='blue',fg='white',font=("arial",14),justify="right")
    totalpasiv_entry.place(x=700,y=720)

    #Button
    generate_button = tk.Button(frame_bs,text="Generate",bg="White",fg="black",font=["arial",12],width=15,command=generate_tb)
    generate_button.place(x=700,y=20)
    balancesheet_button = tk.Button(frame_bs,text="Tampilkan Neraca",bg="White",fg="black",font=["arial",12],width=15,command=balance_sheet)
    balancesheet_button.place(x=700,y=60)
    export_button = tk.Button(frame_bs,text="Export Excel",bg="White",fg="black",font=["arial",12],width=15,command=export_excel)
    export_button.place(x=850,y=60)


    window_bs.mainloop()
 