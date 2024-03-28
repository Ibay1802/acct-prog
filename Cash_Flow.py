import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
import pandas as pd
import openpyxl as ox
from openpyxl import load_workbook
from tkcalendar import Calendar
from tkinter import *
import os
import locale
import datetime
import numpy as np
from num2words import num2words
import math
from pandastable import Table
import sys
from tkinter import scrolledtext
import io
from openpyxl.utils.dataframe import dataframe_to_rows


def cf ():
    window_cf = tk.Tk()
    window_cf.title("Arus Kas")
    window_cf.geometry("1080x800+400+5")
    window_cf.configure(bg='#333333')
    frame_cf=tk.Frame(window_cf,bg="white")
    frame_cf.pack(fill=BOTH,expand=True)

    def start_date(event):
        global cal, date_window
        date_window = Toplevel()
        date_window.grab_set()
        date_window.title('Choose Date')
        date_window.geometry('250x220+400+70')
        cal = Calendar(date_window, selectmode="day", date_pattern="dd/mm/yyyy")
        cal.place(x=0, y=0)
        submit_btn = Button(date_window, text="submit", command=grabstart_date)
        submit_btn.place(x=80, y=190)
    def grabstart_date():
        start_entry.delete(0,END)
        start_entry.insert(0, cal.get_date())
        date_window.destroy()

    def end_date(event):
        global cal, date_window
        date_window = Toplevel()
        date_window.grab_set()
        date_window.title('Choose Date')
        date_window.geometry('250x220+400+70')
        cal = Calendar(date_window, selectmode="day", date_pattern="dd/mm/yyyy")
        cal.place(x=0, y=0)
        submit_btn = Button(date_window, text="submit", command=grab_enddate)
        submit_btn.place(x=80, y=190)
    def grab_enddate():
        end_entry.delete(0, END)
        end_entry.insert(0, cal.get_date())
        date_window.destroy()

    def generate_cf () :
        #try:
            file="database.xlsx"
            df_data = pd.read_excel(file)
            start_date = pd.to_datetime(start_entry.get(),format="%d/%m/%Y", errors="coerce").date()
            end_date = pd.to_datetime(end_entry.get(),format="%d/%m/%Y", errors="coerce").date()
            range_date = pd.date_range(start = start_date, end= end_date)
            data_tanggal = df_data.loc[(df_data[df_data.columns[0]].isin (range_date))]
            selected_cols = data_tanggal[[data_tanggal.columns[0],data_tanggal.columns[1],data_tanggal.columns[9] ,data_tanggal.columns[4], 'Debet', 'Kredit']]
            mask = ~selected_cols['No Akun'].astype(str).str.startswith(('115', '118'))
            mask1 = ~(selected_cols['No Akun'].astype(str).str.startswith('122') & (selected_cols['Kredit'] > 0))
            mask2 = ~selected_cols['Kode Transaksi'].astype(str).isin(['INV', 'JU', 'JP'])
            filtered_cols = selected_cols.loc[mask & mask1 & mask2].dropna()
            print(filtered_cols)
            workbook = ox.load_workbook('lap_keu.xlsx')
            ws = workbook['cash_book']
            for row in dataframe_to_rows(filtered_cols, index=False, header=False):
                ws.append(row)
            workbook.save('lap_keu.xlsx')
            messagebox.showinfo(title="Generate success", message="Generate Success")
        #except:
                #messagebox.showerror("Error",message="Generate Failed")

    df_cash = pd.read_excel('lap_keu.xlsx',sheet_name='cash_book')
    def arus_kas ():
        start_date = pd.to_datetime(start_entry.get(),format="%d/%m/%Y", errors="coerce").date()
        end_date = pd.to_datetime(end_entry.get(),format="%d/%m/%Y", errors="coerce").date()
        range_date = pd.date_range(start = start_date, end= end_date)
        data_tanggal = df_cash.loc[(df_cash[df_cash.columns[0]].isin (range_date))]
        debet_account = data_tanggal.groupby(data_tanggal.columns[2])[["Debet"]].sum().sort_values(by=['Debet'])
        credit_account = data_tanggal.groupby(data_tanggal.columns[2])[["Kredit"]].sum().sort_values(by=['Kredit'])
        saldo_date = df_cash.loc[(pd.to_datetime(df_cash['Tanggal']) < pd.to_datetime(start_date))]
        awal_debit = saldo_date["Debet"].sum()
        awal_credit = saldo_date['Kredit'].sum()
        beg_bal = awal_debit - awal_credit
        #Mapping cash Flow
        saldo_awal = beg_bal
        if 'oprt pay' in debet_account.index:
            cust_inflow = debet_account.loc['oprt pay', 'Debet']
        else:
            cust_inflow = 0
        if 'affiliation' in credit_account.index:
            affiliate_inflow = credit_account.loc['affiliation','Kredit']
        else:
            affiliate_inflow=0
        total_oprinflow = cust_inflow + affiliate_inflow
        if 'cost expense' in debet_account.index:
            cost_outflow = debet_account.loc['cost expense','Debet']
        else:
            cost_outflow = 0
        if 'expense' in debet_account.index:
            adm_outflow = debet_account.loc['expense','Debet']
        else:
            adm_outflow = 0
        debet_tax = 0
        credit_tax = 0
        if 'tax' in debet_account.index:
            debet_tax = debet_account.loc['tax','Debet']
        elif 'tax' in credit_account.index:
            credit_tax = credit_account.loc['tax','Kredit']
        else :
            debet_tax = 0
            credit_tax = 0
        tax_outflow = debet_tax + credit_tax
        total_oproutflow = cost_outflow + adm_outflow + tax_outflow
        net_oprtcashflow = total_oprinflow - total_oproutflow

        if 'investation' in credit_account.index:
            inv_credit = credit_account.loc['investation','Kredit']
        elif 'accm investation' in debet_account.index:
            inv_debet =  debet_account.loc['accm investation','Debet']
        elif 'pl investation' in credit_account.index:
            plinv_credit = credit_account.loc['pl investation','Kredit']
        elif 'pl investation' in debet_account.index:
            plinv_debet = debet_account.loc['pl investation','Debet']
        else :
            inv_credit = 0
            inv_debet = 0
            plinv_credit = 0
            plinv_debet = 0    
        fixasset_inflow = inv_credit - inv_debet + plinv_credit - plinv_debet
        if 'investation' in debet_account.index:
            fixasset_outflow = debet_account.loc['investation','Debet']
        else :
            fixasset_outflow = 0
        invest_outflow = fixasset_outflow
        net_invcashflow = fixasset_inflow-invest_outflow

        if 'bank' in credit_account.index:
            bank_inflow = credit_account.loc['bank','Kredit']
        else :
            bank_inflow = 0
        if 'equity' in credit_account.index:
            equity_inflow = credit_account.loc['equity','Kredit']
        else:
            equity_inflow = 0
        fund_inflow = bank_inflow+equity_inflow
        if 'bank' in debet_account.index:
            bank_outflow = debet_account.get['bank','Debet']
        else:
            bank_outflow = 0
        fund_outflow = bank_outflow
        net_fundcashflow = fund_inflow - fund_outflow

        ending_cashflow = saldo_awal+net_oprtcashflow+net_invcashflow+net_fundcashflow
        
        output_text = scrolledtext.ScrolledText(frame_cf, width=80, height=40)
        output_text.place(x = 20, y=100)

        class TextRedirector(io.TextIOBase):
            def __init__(self, widget):
                self.widget = widget
            def write(self, string):
                self.widget.insert(tk.END, string)
                self.widget.see(tk.END)
                
        sys.stdout = TextRedirector(output_text)
        
        locale.setlocale(locale.LC_ALL, "")
        print(f"                           Laporan Arus Kas")
        print(f"              Per Tanggal laporan {locale.format_string('%s', end_date.strftime('%d/%m/%Y'), grouping=True):>12}")
        print(f"Saldo Kas Awal                 :                  {locale.format_string('%d', saldo_awal, grouping=True):>12}")
        print(f"Arus Kas dari Operasi          :")
        print(f"Ditambah")
        print(f"Kas dari Klien                 : {locale.format_string('%d', cust_inflow, grouping=True):>12}")
        print(f"Kas dari Piutang lainnya       : {locale.format_string('%d', affiliate_inflow, grouping=True):>12}")
        print(f"                                ______________")
        print(f"Total Kas Masuk Operasi        : {locale.format_string('%d',total_oprinflow,grouping=True):>12}")
        print(f"Dikurang")
        print(f"Biaya Usaha                    : {locale.format_string('%d', cost_outflow, grouping=True):>12}")
        print(f"Biaya Administrasi & Umum      : {locale.format_string('%d', adm_outflow , grouping=True):>12}")
        print(f"Pajak                          : {locale.format_string('%d', tax_outflow, grouping=True):>12}")
        print(f"                                ______________")
        print(f"Total Kas Keluar Operasi       : {locale.format_string('%d', total_oproutflow, grouping=True):>12}")
        print(f"                                ______________")
        print(f"Kas Bersih dari Operasi        :                  {locale.format_string('%d', net_oprtcashflow , grouping=True):>12}")
        print(f"Arus Kas Investasi")
        print(f"Ditambah")
        print(f"Kas Masuk Penjualan Aktiva     : {locale.format_string('%d', fixasset_inflow , grouping=True):>12}")
        print(f"                                ______________")
        print(f"Total Kas Masuk Investasi      : {locale.format_string('%d', fixasset_inflow , grouping=True):>12}")
        print(f"Dikurang")
        print(f"Pembelian Aktiva tetap         : {locale.format_string('%d', invest_outflow , grouping=True):>12}")
        print(f"                                ______________")
        print(f"Total Kas keluar Investasi     : {locale.format_string('%d', invest_outflow , grouping=True):>12}")
        print(f"                                ______________")
        print(f"Total Kas Bersih Investasi      :                 {locale.format_string('%d', net_invcashflow , grouping=True):>12}")
        print(f"Arus Kas Pendanaan")
        print(f"Ditambah")
        print(f"Pinjaman Bank                  : {locale.format_string('%d',bank_inflow , grouping=True):>12}")
        print(f"Setoran Modal Saham            : {locale.format_string('%d',equity_inflow , grouping=True):>12}")
        print(f"                                ______________")
        print(f"Total Kas Masuk Pendanaan      : {locale.format_string('%d', fund_inflow , grouping=True):>12}")
        print(f"Dikurang")
        print(f"Biaya Bank                     : {locale.format_string('%d',fund_outflow , grouping=True):>12}")
        print(f"                                ______________")
        print(f"Total Kas keluar Pendanaan     : {locale.format_string('%d', fund_outflow, grouping=True):>12}")
        print(f"                                ______________")
        print(f"Total Kas Bersih Pendanaan     :                  {locale.format_string('%d', net_fundcashflow, grouping=True):>12}")
        print(f"                                ______________")
        print(f"Saldo Akhir Kas                :                  {locale.format_string('%d', ending_cashflow, grouping=True):>12}")
        print(f"                                ==========================================")

    sys.stdout = sys.__stdout__

    template_cf = ox.load_workbook("template_CF.xlsx")
    ws = template_cf["Cash Flow"]
    def print_excel ():
        start_date = pd.to_datetime(start_entry.get(),format="%d/%m/%Y", errors="coerce").date()
        end_date = pd.to_datetime(end_entry.get(),format="%d/%m/%Y", errors="coerce").date()
        range_date = pd.date_range(start = start_date, end= end_date)
        data_tanggal = df_cash.loc[(df_cash[df_cash.columns[0]].isin (range_date))]
        debet_account = data_tanggal.groupby(data_tanggal.columns[2])[["Debet"]].sum().sort_values(by=['Debet'])
        credit_account = data_tanggal.groupby(data_tanggal.columns[2])[["Kredit"]].sum().sort_values(by=['Kredit'])
        saldo_date = df_cash.loc[(pd.to_datetime(df_cash['Tanggal']) < pd.to_datetime(start_date))]
        awal_debit = saldo_date["Debet"].sum()
        awal_credit = saldo_date['Kredit'].sum()
        beg_bal = awal_debit - awal_credit
        #Mapping cash Flow
        saldo_awal = beg_bal
        if 'oprt pay' in debet_account.index:
            cust_inflow = debet_account.loc['oprt pay', 'Debet']
        else:
            cust_inflow = 0
        if 'affiliation' in credit_account.index:
            affiliate_inflow = credit_account.loc['affiliation','Kredit']
        else:
            affiliate_inflow=0
        total_oprinflow = cust_inflow + affiliate_inflow
        if 'cost expense' in debet_account.index:
            cost_outflow = debet_account.loc['cost expense','Debet']
        else:
            cost_outflow = 0
        if 'expense' in debet_account.index:
            adm_outflow = debet_account.loc['expense','Debet']
        else:
            adm_outflow = 0
        debet_tax = 0
        credit_tax = 0
        if 'tax' in debet_account.index:
            debet_tax = debet_account.loc['tax','Debet']
        elif 'tax' in credit_account.index:
            credit_tax = credit_account.loc['tax','Kredit']
        else :
            debet_tax = 0
            credit_tax = 0
        tax_outflow = debet_tax + credit_tax
        total_oproutflow = cost_outflow + adm_outflow + tax_outflow
        net_oprtcashflow = total_oprinflow - total_oproutflow

        if 'investation' in credit_account.index:
            inv_credit = credit_account.loc['investation','Kredit']
        elif 'accm investation' in debet_account.index:
            inv_debet =  debet_account.loc['accm investation','Debet']
        elif 'pl investation' in credit_account.index:
            plinv_credit = credit_account.loc['pl investation','Kredit']
        elif 'pl investation' in debet_account.index:
            plinv_debet = debet_account.loc['pl investation','Debet']
        else :
            inv_credit = 0
            inv_debet = 0
            plinv_credit = 0
            plinv_debet = 0    
        fixasset_inflow = inv_credit - inv_debet + plinv_credit - plinv_debet
        if 'investation' in debet_account.index:
            fixasset_outflow = debet_account.loc['investation','Debet']
        else :
            fixasset_outflow = 0
        invest_outflow = fixasset_outflow
        net_invcashflow = fixasset_inflow-invest_outflow

        if 'bank' in credit_account.index:
            bank_inflow = credit_account.loc['bank','Kredit']
        else :
            bank_inflow = 0
        if 'equity' in credit_account.index:
            equity_inflow = credit_account.loc['equity','Kredit']
        else:
            equity_inflow = 0
        fund_inflow = bank_inflow+equity_inflow
        if 'bank' in debet_account.index:
            bank_outflow = debet_account.get['bank','Debet']
        else:
            bank_outflow = 0
        fund_outflow = bank_outflow
        net_fundcashflow = fund_inflow - fund_outflow
        ws['A4'] = "Per Tanggal laporan "+end_entry.get()
        ws['D5'] = int(saldo_awal)
        ws['C8'] = int(cust_inflow)
        ws['C9'] = int(affiliate_inflow)
        ws['C12'] = int(cost_outflow)
        ws['C13'] = int(adm_outflow)
        ws['c14'] = int(tax_outflow)
        ws['c19'] = int(fixasset_inflow)
        ws['c22'] = int(invest_outflow)
        ws['c27'] = int(bank_inflow)
        ws['c28'] = int(equity_inflow)
        ws['c31'] = int(fund_outflow)

        template_cf.save('template_CF.xlsx')
        os.startfile('template_CF.xlsx')
        messagebox.showinfo(title="Export Cash Flow", message="Export Lap. Cash Flow Success")




    cf_label = tk.Label(frame_cf,text="Laporan Arus Kas",bg="White",fg="black",font=["arial",20],justify="center")
    cf_label.place(x=200,y=10)
    per_label=tk.Label(frame_cf,text=" - ",bg="white",fg="black",font=("arial",12))
    per_label.place(x=798,y=10)
    start_entry = tk.Entry(frame_cf, highlightthickness=0,bg="white",fg="black",font=("arial",12),width=10)
    start_entry.place(x=700, y=10)
    start_entry.insert(0, "dari tanggal")
    start_entry.bind("<1>",start_date)
    end_entry = tk.Entry(frame_cf, highlightthickness=0,bg="white",fg="black",font=("arial",12),width=12)
    end_entry.place(x=820, y=10)
    end_entry.insert(0, "sampai tanggal")
    end_entry.bind("<1>",end_date)
    end_entry.bind("<FocusOut>", lambda event:until_entry.delete(0, END) or  until_entry.insert(0,end_entry.get()))
    until_label = tk.Label(frame_cf,text="sampai dengan tanggal",bg="White",fg="black",font=["arial",12],justify="center")
    until_label.place(x=180,y=58)
    until_entry = tk.Entry(frame_cf,bg="white",fg="black",font=("arial",10),width=10)
    until_entry.place(x=350, y=60)






    #Button
    generate_button = tk.Button(frame_cf,text="Generate",bg="grey",fg="white",font=["arial",12],width=25,command=generate_cf)
    generate_button.place(x=700,y=40)
    cash_flow_button = tk.Button(frame_cf,text="Tampilkan Cash Flow",bg="grey",fg="white",font=["arial",12],width=25,command=arus_kas)
    cash_flow_button.place(x=700,y=90)
    print_button = tk.Button(frame_cf,text="Print Cash Flow",bg="black",fg="white",font=["arial",12],width=25,command=print_excel)
    print_button.place(x=700,y=140)

    window_cf.mainloop()