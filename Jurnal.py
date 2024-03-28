import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
import pandas as pd
import openpyxl as ox
from openpyxl import load_workbook
from tkinter import *
from tkcalendar import Calendar
import os
import locale
import datetime
import numpy as np
from num2words import num2words
import math
from pandastable import Table
import sys


#Jurnal
def jurnal ():
    window_jurnal = tk.Tk()
    window_jurnal.title("Jurnal")
    window_jurnal.geometry("1080x780+400+10")
    window_jurnal.configure(bg='#333333')
    frame_jurnal=tk.Frame(window_jurnal,bg='#333333')
    frame_jurnal.pack(fill=BOTH,expand=True)

    #Kalender
    def pick_date(event):
        global cal, date_window
        date_window = Toplevel()
        date_window.grab_set()
        date_window.title('Choose Date')
        date_window.geometry('250x220+590+370')
        cal = Calendar(date_window, selectmode="day", date_pattern="dd/mm/yyyy")
        cal.place(x=0, y=0)
        submit_btn = Button(date_window, text="submit", command=grab_date)
        submit_btn.place(x=80, y=190)
    def grab_date():
        tanggal_entry.delete(0,END)
        tanggal_entry.insert(0, cal.get_date())
        date_window.destroy()
    # nama akun vlookup dr no akun
    def vlookup(value):
        filter_value = combobox_var.get()
        noakun_menu['values'] = [value for value in noakun_menu['values'] if value.startswith(filter_value)]
        result = data[data[data.columns[0]] == value][data.columns[1]].values[0]
        return result

    #record baris jurnal       
    jurnal_list = []
    def rekam_baris():
        No_Akun = noakun_menu.get()
        Nama_Akun = vlookup(No_Akun)
        Debet = int(debet_entry.get())
        Kredit = int(kredit_entry.get())
        Keterangan = keterangan_entry.get()
        if No_Akun[:3] in(['501', '505', '506', '214', '216']):
            category = "cost expense"
        elif No_Akun[:3] in(['610', '117', '211', '210']):
            category = "expense"
        elif No_Akun[:3] in(['116', '219']):
            category = "affiliation"
        elif No_Akun[:3] in(['118', '212', '750']):
            category = "tax"
        elif No_Akun[:3] in([ '114','213', '220']):
            category = "bank"
        elif No_Akun in(['730-0001', '730-0002', '730-0003', '730-0004','710-0001','710-0005']):
            category = "bank"
        elif No_Akun[:3] in(['121']):
            category = "investation"
        elif No_Akun in(['730-0006,710-0006']):
            category = "pl investation"
        elif No_Akun[:3] in(['122']):
            category = "accm investation"
        elif No_Akun[:3] in(['310']):
            category = "equity" 
        else:
            category = ""

        jurnal_item = [No_Akun, Nama_Akun, Debet, Kredit, Keterangan,category]
        tree.insert('',index= 'end', values=jurnal_item)
        noakun_menu.delete(0, END)
        namaakun_entry.delete(0,END)
        debet_entry.delete(0, END)
        kredit_entry.delete(0,END)
        keterangan_entry.delete(0,END)
        jurnal_list.append(jurnal_item)
    #hapus jurnal salah
    def hapus_terpilih():
        x = tree.selection()
        for record in x:
            tree.delete(record)
    #New Jurnal
    def jurnal_baru ():
        confirmation1 = messagebox.askyesno("confirmation","apakah data sudah di generate?")
        if confirmation1 :
            tanggal_entry.delete(0,END)
            kodeTransaksi_entry.delete(0,END)
            noTransaksi_entry.delete(0,END)
            transaksi_entry.delete(0,END)
            noakun_menu.delete(0, END)
            namaakun_entry.delete(0,END)
            debet_entry.delete(0, END)
            kredit_entry.delete(0,END)
            keterangan_entry.delete(0,END)
            debet_sum.delete(0,END)
            kredit_sum.delete(0,END)
            tree.delete(*tree.get_children())
            jurnal_list.clear()
        else:
            confirmation2 = messagebox.askyesno("confirmation","apakan anda yakin akan membuat jurnal baru?")
            if confirmation2:
                tanggal_entry.delete(0,END)
                kodeTransaksi_entry.delete(0,END)
                noTransaksi_entry.delete(0,END)
                transaksi_entry.delete(0,END)
                noakun_menu.delete(0, END)
                namaakun_entry.delete(0,END)
                debet_entry.delete(0, END)
                kredit_entry.delete(0,END)
                keterangan_entry.delete(0,END)
                debet_sum.delete(0,END)
                kredit_sum.delete(0,END)
                tree.delete(*tree.get_children())
                jurnal_list.clear()
            else :
                return
    def new_jurnal ():
        tanggal_entry.delete(0,END)
        kodeTransaksi_entry.delete(0,END)
        noTransaksi_entry.delete(0,END)
        transaksi_entry.delete(0,END)
        noakun_menu.delete(0, END)
        namaakun_entry.delete(0,END)
        debet_entry.delete(0, END)
        kredit_entry.delete(0,END)
        keterangan_entry.delete(0,END)
        debet_sum.delete(0,END)
        kredit_sum.delete(0,END)
        tree.delete(*tree.get_children())
        jurnal_list.clear()


    #Generate Jurnal
    def generate ():
        #date = tanggal_entry.get()
        #year = date[-4:]
        #filepath=f'database_{year}.xlsx'
        #if not os.path.exists(filepath):    
        #    workbook = ox.Workbook()
        #    ws = workbook.active
        #    heading = ["Tanggal,Kode Transaksi,No Transaksi,Transaksi,No Akun,Nama Akun,Debet,Kredit,Keterangan"]
        #    ws.append(heading)
        #    workbook.save(filepath)
        filepath = "database.xlsx"
        workbook = ox.load_workbook(filepath)
        ws = workbook.active
        Tanggal = tanggal_entry.get()
        Kode_transaksi =  kodeTransaksi_entry.get()
        No_Transaksi = noTransaksi_entry.get()
        Transaksi = transaksi_entry.get()
        for jurnal_item in tree.get_children():
            values = tree.item(jurnal_item)["values"]
            ws.append([Tanggal,Kode_transaksi,No_Transaksi,Transaksi] + values)
        workbook.save(filepath)
        new_jurnal ()

    #Design Input Jurnal    
    nama_label = tk.Label(frame_jurnal, text="PT. ABCDEFU",bg='#333333',fg='#FFFFFF',font=["arial",20],justify="center")
    nama_label.grid(row=0, column=0, columnspan=3,padx=40)
    form_label = tk.Label(frame_jurnal, text="Form Jurnal",bg='#333333',fg='white',font=["arial",20],justify='center')
    form_label.grid(row=1,column=0,columnspan=3,pady=10,padx=10)
    #tanggal
    tanggal_label = tk.Label(frame_jurnal, text="Tanggal",bg ='#333333', fg='white',font=["arial",14])
    tanggal_label.place(x=20, y=150)
    tanggal_entry = tk.Entry(frame_jurnal, highlightthickness=0, relief=FLAT,bg="white",fg="#6b6a69",font=("arial",14))
    tanggal_entry.place(x=150, y=150)
    tanggal_entry.insert(0, "dd/mm/yyyy")
    tanggal_entry.bind("<1>",pick_date)
    #No. Transaksi
    noTransaksi_label = tk.Label(frame_jurnal, text="No. Transaksi", bg='#333333', fg='white', font=["arial",14])
    noTransaksi_label.place(x=20, y=200)
    options = ["CI","CO","JU","JP"]
    kodeTransaksi_entry = ttk.Combobox(frame_jurnal,values=options, width=3,font=["arial",12])
    kodeTransaksi_entry.place(x=150,y=200)
    noTransaksi_entry = tk.Entry(frame_jurnal, width=18,highlightthickness=0, relief=FLAT,bg="white",fg="#6b6a69",font=("arial",13))
    noTransaksi_entry.place(x=205,y=200)
    #Keterangan
    transaksi_label = tk.Label(frame_jurnal, text ="Transaksi", bg='#333333', fg='white', font=["arial",14])
    transaksi_label.place(x=20, y=250)
    transaksi_entry = tk.Entry(frame_jurnal,width=80,bg="white",fg="#6b6a69",font=("arial",14))
    transaksi_entry.place(x=150,y=250)
    #Table Jurnal
    noakun_label = tk.Label(frame_jurnal, text = "No. Akun",bg='#333333', fg='white', font=["arial",14])
    noakun_label.place(x=20,y=300)
    #menentukan nama akun dan no akun
    data = pd.read_excel("lap_keu.xlsx", sheet_name="COA")
    #data option no.akun
    options = data[data.columns[0]].tolist()
    combobox_var = tk.StringVar()
    #Design Layout
    noakun_menu = ttk.Combobox(frame_jurnal, values=options,width=15,textvariable=combobox_var)
    noakun_menu.place(x=20, y=330)
    noakun_menu.bind("<<ComboboxSelected>>", lambda event:namaakun_entry.delete(0, END) or  namaakun_entry.insert(0, vlookup(noakun_menu.get())))
    #Nama Akun
    namaakun_label = tk.Label(frame_jurnal,text="Nama Akun",bg='#333333', fg='white', font=["arial",12])
    namaakun_label.place(x=160,y=300)
    #setting nama akun vlookup no.akun
    namaakun_entry = tk.Entry(frame_jurnal, width=25 ,bg="white",fg="black",font=("arial",12))
    namaakun_entry.place(x=160,y=330)

    #debet
    debet_label=tk.Label(frame_jurnal,text="Debet",bg='#333333', fg='white', font=["arial",12])
    debet_label.place(x=430, y=300)
    debet_entry = tk.Entry(frame_jurnal,width=15 ,exportselection=0,bg="white",fg="black", font=["arial",12])
    debet_entry.place(x=430, y=330)

    #Kredit
    kredit_label=tk.Label(frame_jurnal,text="Kredit",bg='#333333', fg='white', font=["arial",12])
    kredit_label.place(x=600, y=300)
    kredit_entry = tk.Entry(frame_jurnal,width=15 ,exportselection=0,bg="white",fg="black", font=["arial",12])
    kredit_entry.place(x=600, y=330)

    #Keterangan
    keterangan_label = tk.Label(frame_jurnal, text ="Keterangan", bg='#333333', fg='white', font=["arial",12])
    keterangan_label.place(x=750, y=300)
    keterangan_entry = tk.Entry(frame_jurnal,width=35,bg="white",fg="black",font=("arial",12))
    keterangan_entry.place(x=750,y=330)

    #Display Jurnal table
    columns = ('no_akun', 'nama_akun', 'debet', 'kredit','keterangan','category')
    tree = ttk.Treeview(frame_jurnal, columns=columns, show="headings")
    tree.heading('no_akun', text="No. Akun")
    tree.heading('nama_akun', text="Nama Akun")
    tree.heading('debet', text="Debet")
    tree.heading('kredit', text="Kredit")
    tree.heading('keterangan', text="Keterangan")
    tree.heading('category', text="kategori")
    tree.place(x=20, y=400)


    debet_sum = ttk.Entry(frame_jurnal)
    debet_sum.place(x=430,y=650)
    kredit_sum = ttk.Entry(frame_jurnal)
    kredit_sum.place(x=600,y=650)
    balance_check = ttk.Label(frame_jurnal)
    balance_check.place(x= 750, y=650)

    def update_sum():
        column_debet = [int(tree.set(child,'debet'))for child in tree.get_children('')]
        sum_debet = sum(column_debet)
        debet_sum.delete(0, 'end')
        debet_sum.insert(0, sum_debet)
        column_kredit = [int(tree.set(child,'kredit'))for child in tree.get_children('')]
        sum_kredit = sum(column_kredit)
        kredit_sum.delete(0, 'end')
        kredit_sum.insert(0, sum_kredit)
        if sum_debet-sum_kredit == 0 :
                balance_check.config(text="BALANCE",background='green')
        else :
                balance_check.config(text="UNBALANCE",background='red' )

    update_sum()

    #Button add
    add_akun_button = tk.Button(frame_jurnal, text = "Masukan Jurnal",command=rekam_baris)
    add_akun_button.place(x=750, y=360)
    erase_akun_button = tk.Button(frame_jurnal, text = "Hapus Baris",command=hapus_terpilih)
    erase_akun_button.place(x=850, y=360)
    check_button = tk.Button(frame_jurnal, text="Check", command= update_sum)
    check_button.place(x= 850, y=650)

    #Generate jurnal button
    generate_jurnal = tk.Button(frame_jurnal, text = "Generate Jurnal",command=generate)
    generate_jurnal.grid(row=6, column=0, columnspan=15, sticky="news", padx=20, pady=600)
    #Jurnal Baru button
    jurnal_baru = tk.Button(frame_jurnal,text="Jurnal Baru",command=jurnal_baru)
    jurnal_baru.place(x=20,y=730, width=500)

    window_jurnal.mainloop