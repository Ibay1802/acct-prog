import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
import pandas as pd
import openpyxl as ox
from openpyxl import load_workbook
from tkinter import *
from tkcalendar import Calendar
import os
import locale
import datetime
import numpy as np
from num2words import num2words
import math
from pandastable import Table
import sys

# Invoice
def invoice ():
    window_invoice = tk.Tk()
    window_invoice.title("Invoice_Entry")
    window_invoice.geometry("1080x780+400+10")
    window_invoice.configure(bg='#333333')
    frame_inv=tk.Frame(window_invoice,bg='#333333')
    frame_inv.pack(fill=BOTH,expand=True)

    def pick_date(event):
        global cal, date_window
        date_window = Toplevel()
        date_window.grab_set()
        date_window.title('Choose Date')
        date_window.geometry('250x220+590+370')
        cal = Calendar(date_window, selectmode="day", date_pattern="dd/mm/yyyy")
        cal.place(x=0, y=0)
        submit_btn = Button(date_window, text="submit", command=grab_date)
        submit_btn.place(x=80, y=190)
    def pick_duedate(event):
        global cal, date_window
        date_window = Toplevel()
        date_window.grab_set()
        date_window.title('Choose Date')
        date_window.geometry('250x220+590+370')
        cal = Calendar(date_window, selectmode="day", date_pattern="dd/mm/yyyy")
        cal.place(x=0, y=0)
        submit_btn = Button(date_window, text="submit", command=grab_duedate)
        submit_btn.place(x=80, y=190)
    def grab_date():
        tanggal_entry.delete(0,END)
        tanggal_entry.insert(0, cal.get_date())
        date_window.destroy()
    def grab_duedate():
        duedate_entry.delete(0,END)
        duedate_entry.insert(0, cal.get_date())
        date_window.destroy()   

    #vlookup nama cust
    def vlookup(value):
        filter_value = combobox_var.get()
        cust_kode['values'] = [value for value in cust_kode['values'] if isinstance(value, str) and value.startswith(filter_value)]
        result = custfile[custfile[custfile.columns[0]] == value][custfile.columns[1]].values[0]
        return result        
    #vlookup alamat cust
    def addlookup(entry_value):
        filtered_values = [value for value in custfile[custfile.columns[1]].values if isinstance(value, str) and value.startswith(entry_value)]
        if filtered_values:
            result1 = custfile[custfile[custfile.columns[1]] == filtered_values[0]][custfile.columns[2]].values[0]
        else:
            result1 = ""
        return result1
    #vlookup nama akun
    def aklookup(value):
        filter_value1 = combobox_var.get()
        akun_entry['values'] = [val for val in akun_entry['values'] if isinstance(val, str) and val.startswith(filter_value1)]
        result_series = df.loc[df['noakun'] == value, 'namaakun']
        if not result_series.empty:
            result2 = result_series.iloc[0]
        else:
            result2 = 'No match found'
        return result2

    #record baris jurnal       
    def rekam_baris():
        No_Akun =akun_entry.get()
        No_Akun1 = akun_entry1.get()
        if No_Akun in "401-0001":
                    No_Akun2 = "115-0001"
        elif No_Akun == "401-0002":
                    No_Akun2 = "115-0002"
        elif No_Akun == "401-0003":
                    No_Akun2 = "115-0003"
        elif No_Akun == "401-0099":
                    No_Akun2 = "115-0099"
        else :
                    No_Akun2 = "default value"
        Nama_Akun = aklookup(No_Akun)
        Nama_Akun1 = aklookup(No_Akun1)
        Nama_Akun2 = aklookup(No_Akun2)
        Debet = int(total_entry.get())
        Kredit1 = int(jumlah_entry.get())
        Kredit2 = int(ppn_entry.get()) 
        Keterangan = uraian_entry.get()
        row_data = [(No_Akun2, Nama_Akun2, Debet, "0", Keterangan),
                    (No_Akun, Nama_Akun, "0", Kredit1, Keterangan),
                    (No_Akun1, Nama_Akun1, "0", Kredit2, Keterangan)]
        for data in row_data:
            tree.insert("", tk.END, values=data)

    #Jurnal Baru
    def new_inv ():
        inv_entry.delete(0,END)
        tanggal_entry.delete(0,END)
        duedate_entry.delete(0,END)
        cust_kode.delete(0,END)
        custname_entry.delete(0,END)
        alamat_cust.delete(0, END)
        akun_entry.delete(0,END)
        akun_entry1.delete(0,END)
        nama_akun.delete(0,END)
        nama_akun1.delete(0,END)
        uraian_entry.delete(0,END)
        jumlah_entry.delete(0,END)
        ppn_entry.delete(0,END)
        total_entry.delete(0,END)
        terbilang_entry.delete(0,END)
        tree.delete(*tree.get_children())

    def inv_baru ():
        confirmation1 = messagebox.askyesno("confirmation","apakah data sudah di generate?")
        if confirmation1 :
            inv_entry.delete(0,END)
            tanggal_entry.delete(0,END)
            duedate_entry.delete(0,END)
            cust_kode.delete(0,END)
            custname_entry.delete(0,END)
            alamat_cust.delete(0, END)
            akun_entry.delete(0,END)
            akun_entry1.delete(0,END)
            nama_akun.delete(0,END)
            nama_akun1.delete(0,END)
            uraian_entry.delete(0,END)
            jumlah_entry.delete(0,END)
            ppn_entry.delete(0,END)
            total_entry.delete(0,END)
            terbilang_entry.delete(0,END)
            tree.delete(*tree.get_children())
        else:
            confirmation2 = messagebox.askyesno("confirmation","apakan anda yakin akan membuat invoice baru?")
            if confirmation2:
                inv_entry.delete(0,END)
                tanggal_entry.delete(0,END)
                duedate_entry.delete(0,END)
                cust_kode.delete(0,END)
                custname_entry.delete(0,END)
                alamat_cust.delete(0, END)
                akun_entry.delete(0,END)
                akun_entry1.delete(0,END)
                nama_akun.delete(0,END)
                nama_akun1.delete(0,END)
                uraian_entry.delete(0,END)
                jumlah_entry.delete(0,END)
                ppn_entry.delete(0,END)
                total_entry.delete(0,END)
                terbilang_entry.delete(0,END)
                tree.delete(*tree.get_children())
            else :
                return


    file2= ox.load_workbook('Cust_file.xlsx')
    ws2 = file2['Data_Base_Invoice']
    def inv_generate ():  
        #date = tanggal_entry.get()
        #year = date[-4:]
        #file1= f'database_{year}.xlsx'
        #if not os.path.exists(file1):    
        #    workbook = ox.Workbook()
        #    ws = workbook.active
        #    heading = ["Tanggal,Kode Transaksi,No Transaksi,Transaksi,No Akun,Nama Akun,Debet,Kredit,Keterangan"]
        #    ws.append(heading)
         #   workbook.save(file1)
        file1 = "database.xlsx"
        workbook = ox.load_workbook(file1)
        ws = workbook.active
        nomor_inv = inv_entry.get()
        inv_no = "INV" + inv_entry.get()
        tanggal_inv = tanggal_entry.get()
        duedate_inv = duedate_entry.get()
        kode_cust = cust_kode.get()
        nama_cust = custname_entry.get()
        alamat = alamat_cust.get()
        uraian = uraian_entry.get()
        jumlah = int(jumlah_entry.get())
        ppn = int(ppn_entry.get())
        total = int(total_entry.get())
        terbilang = terbilang_entry.get()
        keterangan = uraian_entry.get() + "INV" + inv_entry.get()
        No_Akun = akun_entry.get()
        if No_Akun in "401-0001":
                    No_Akun2 = "115-0001"
        elif No_Akun == "401-0002":
                    No_Akun2 = "115-0002"
        elif No_Akun == "401-0003":
                    No_Akun2 = "115-0003"
        elif No_Akun == "401-0099":
                    No_Akun2 = "115-0099"
        else :
                    No_Akun2 = "default value"
        Nama_Akun2 = aklookup(No_Akun2)
        row_values = []
        for row_data in tree.get_children():
            values = tree.item(row_data)["values"]
            row_values.append(values)           
        for row in row_values:
            ws.append([tanggal_inv, "INV", nomor_inv, keterangan] + row) #export ke database jurnal
        #export ke database invoice
        ws2.append([inv_no,kode_cust,nama_cust,alamat,tanggal_inv,duedate_inv,uraian,jumlah,ppn,total,No_Akun2,Nama_Akun2,terbilang])
        workbook.save(file1)
        file2.save('Cust_file.xlsx')
        
        new_inv()

    template_invoice = ox.load_workbook('forminvoice.xlsx')
    sheet = template_invoice['Form Invoice']
    def print_inv():
        inv_no = "INV" + inv_entry.get()
        name = custname_entry.get()
        sheet['i13'] = inv_no
        sheet['i14'] = tanggal_entry.get()
        sheet['i15'] = duedate_entry.get()
        sheet['b17'] = name
        sheet['b18'] =alamat_cust.get()
        sheet['d25'] =uraian_entry.get()
        sheet['d34'] = terbilang_entry.get()
        sheet['h25'] =int(jumlah_entry.get())
        new_file = inv_no + name + " " + datetime.datetime.now().strftime("%Y-%m-%d")+".xlsx"
        template_invoice.save(new_file)
        os.startfile(new_file)
        
    #Design Invoice
    inv_label = tk.Label(frame_inv,text="INVOICE",bg='#333333',fg='#FFFFFF',font=["arial",20],justify="center")
    inv_label.pack(side=TOP)
    #Nomor Invoice
    no_inv = tk.Label(frame_inv,text="Nomor Invoice",bg='#333333',fg='#FFFFFF',font=["arial",14],justify="left")
    no_inv.place(x=20,y=70)
    inv_kode = tk.Label(frame_inv, text = "INV", bg="white", fg="black",font=["arial",12])
    inv_kode.place(x=180,y=70)
    inv_entry = tk.Entry(frame_inv, bg="white",fg="black",font=["arial",12],width=20)
    inv_entry.place(x=220,y=70)
    #tanggal
    tanggal_label = tk.Label(frame_inv, text="Tanggal",bg ='#333333', fg='white',font=["arial",14])
    tanggal_label.place(x=20, y=100)
    tanggal_entry = tk.Entry(frame_inv, highlightthickness=0, relief=FLAT,bg="white",fg="#6b6a69",font=("arial",14))
    tanggal_entry.place(x=180, y=100)
    tanggal_entry.insert(0, "dd/mm/yyyy")
    tanggal_entry.bind("<1>",pick_date)
    duedate_label=tk.Label(frame_inv,text="Jatuh Tempo",bg='#333333', fg='white',font=["arial",14])
    duedate_label.place(x=400, y=100)
    duedate_entry = tk.Entry(frame_inv, highlightthickness=0, relief=FLAT,bg="white",fg="#6b6a69",font=("arial",14))
    duedate_entry.place(x=550, y=100)
    duedate_entry.insert(0, "dd/mm/yyyy")
    duedate_entry.bind("<1>",pick_duedate)
    #Customer
    custfile = pd.read_excel("Cust_file.xlsx", sheet_name="Data Base Cust")
    options = custfile[custfile.columns[0]].tolist()
    combobox_var = tk.StringVar()
    text_var = tk.StringVar()
    cust_label = tk.Label(frame_inv,text="Customer",bg ='#333333', fg='white',font=["arial",14])
    cust_label.place(x=20, y=130)
    cust_kode = ttk.Combobox(frame_inv,values=options,width=10,textvariable=combobox_var)
    cust_kode.place(x=30, y=160)
    cust_kode.bind("<<ComboboxSelected>>", lambda event:custname_entry.delete(0, END) or  custname_entry.insert(0, vlookup(cust_kode.get())))
    custname_entry = tk.Entry(frame_inv, bg="white",fg="black",font=["arial",12],width=40)
    custname_entry.place(x=30,y=190)
    custname_entry.bind("<FocusOut>", lambda event: alamat_cust.delete(0, tk.END) or alamat_cust.insert(0, addlookup(event.widget.get())))
    alamat_cust = tk.Entry(frame_inv, bg="white",fg="black",font=["arial",12],width=80)
    alamat_cust.place(x=30,y=220)
    akun_penjualan = tk.Label(frame_inv,text="Akun Penjualan",bg ='#333333', fg='white',font=["arial",14])
    akun_penjualan.place(x=30, y=260)

    def update_ppn_entry(event=None):
        jumlah = int(jumlah_entry.get())
        ppn_value = int(jumlah * 0.11)
        ppn_entry.delete(0, END)
        ppn_entry.insert(0, ppn_value)

    def update_total(event=None):
        jumlah = int(jumlah_entry.get() or 0)
        ppn = (jumlah*0.11)
        total = int(jumlah + ppn)
        total_entry.delete(0, END)
        total_entry.insert(0, total)
        terbilang_value = get_terbilang(total)
        terbilang_entry.delete(0, tk.END)
        terbilang_entry.insert(0, terbilang_value + " Rupiah")

    def get_terbilang(angka):
        # Definisikan fungsi terbilang di sini
        satuan = ["", "satu", "dua", "tiga", "empat", "lima", "enam", "tujuh", "delapan", "sembilan"]
        belasan = ["", "sebelas", "dua belas", "tiga belas", "empat belas", "lima belas", "enam belas", "tujuh belas", "delapan belas", "sembilan belas"]
        puluhan = ["", "sepuluh", "dua puluh", "tiga puluh", "empat puluh", "lima puluh", "enam puluh", "tujuh puluh", "delapan puluh", "sembilan puluh"]

        terbilang_result = ""

        # Miliar
        if angka // 1000000000 > 0:
            terbilang_result += get_terbilang(angka // 1000000000) + " miliar "
            angka %= 1000000000

        # Juta
        if angka // 1000000 > 0:
            terbilang_result += get_terbilang(angka // 1000000) + " juta "
            angka %= 1000000

        # Ribu
        if angka // 1000 > 0:
            terbilang_result += get_terbilang(angka // 1000) + " ribu "
            angka %= 1000

        # Ratusan
        if angka // 100 > 0:
            if angka // 100 == 1:
                terbilang_result += "seratus "
            else:
                terbilang_result += satuan[angka // 100] + " ratus "
            angka %= 100

        # Belasan dan puluhan
        if angka >= 20:
            terbilang_result += puluhan[angka // 10] + " "
            angka %= 10
        elif angka >= 11:
            terbilang_result += belasan[angka - 10] + " "
            angka = 0

        # Satuan
        if angka > 0:
            terbilang_result += satuan[angka] + " "

        terbilang_result += " "
        return terbilang_result.strip()

    akun_no = ["401-0001","401-0002","401-0003","401-0099","212-0001"]
    akun ={'noakun': ["401-0001","401-0002","401-0003","401-0099","212-0001","115-0001","115-0002","115-0003","115-0099"],
            'namaakun':["Pendapatan Usaha Projek","Pendapatan Usaha Training","Pendapatan Usaha Teknikal Service","Pendapatan Usaha Lainnya","Hutang Pajak PPN Keluaran","Piutang Project","Piutang Training","Piutang Teknikal Service","Piutang usaha lainnya"]}
    df= pd.DataFrame(akun)
    akun_entry = ttk.Combobox(frame_inv, values=akun_no, width=15)
    akun_entry.place(x=30, y=290)
    akun_entry.bind("<<ComboboxSelected>>", lambda event: nama_akun.delete(0, END) or nama_akun.insert(0, aklookup(akun_entry.get())))
    akun_entry1 = ttk.Combobox(frame_inv, values="212-0001", width=15)
    akun_entry1.place(x=30, y=320)
    akun_entry1.bind("<<ComboboxSelected>>", lambda event: nama_akun1.delete(0, END) or nama_akun1.insert(0, aklookup(akun_entry1.get())))
    namaakun_label = tk.Label(frame_inv,text="Nama Akun",bg ='#333333', fg='white',font=["arial",14])
    namaakun_label.place(x=180, y=260)
    nama_akun = tk.Entry(frame_inv,bg="white",fg="black",font=["arial",12])
    nama_akun.place(x=180,y=290)
    nama_akun1 = tk.Entry(frame_inv,bg="white",fg="black",font=["arial",12])
    nama_akun1.place(x=180,y=320)

    uraian_label = tk.Label(frame_inv,text="Uraian",bg ='#333333', fg='white',font=["arial",14])
    uraian_label.place(x=430, y= 260)
    uraian_entry = tk.Entry(frame_inv,width=40,bg="white",fg="black",font=["arial",12])
    uraian_entry.place(x=430,y=290)
    jumlah_label = tk.Label(frame_inv,text= "Jumlah",bg ='#333333', fg='white',font=["arial",14])
    jumlah_label.place(x=850, y=260)     
    jumlah_entry = tk.Entry(frame_inv, bg="white", fg="black", font=["arial", 12])
    jumlah_entry.place(x=850, y=290)

    ppn_label = tk.Label(frame_inv, text = "PPN", bg="#333333", fg="white", font=["arial", 12])
    ppn_label.place(x=420,y=320)
    ppn_entry = tk.Entry(frame_inv,bg="white", fg="black", font=["arial", 12])
    ppn_entry.place(x=850, y=320)
    ppn_entry.bind("<Return>", update_ppn_entry)
    total_label = tk.Label(frame_inv, text="Total Jumlah", bg="#333333", fg="white", font=["arial", 14])
    total_label.place(x=420, y=350)
    total_entry = tk.Entry(frame_inv, bg="white", fg="black", font=["arial", 12])
    total_entry.place(x=850, y=350)
    total_entry.bind("<Return>", lambda event: update_total())

    terbilang_label = tk.Label(frame_inv, text="Total Terbilang", bg="#333333", fg="white", font=["arial", 14])
    terbilang_label.place(x=30, y=360)
    terbilang_entry = tk.Entry(frame_inv, bg="white",fg="black", font=["arial", 12], width=85)
    terbilang_entry.place(x=30, y=390)
    total_entry.bind("<Return>", lambda event: update_total())
    update_total()

    #Pembuatan Jurnal Invoice
    #Display Jurnal Invoice
    columns = ('no_akun', 'nama_akun', 'debet', 'kredit','keterangan')
    tree = ttk.Treeview(frame_inv, columns=columns, show="headings")
    tree.heading('no_akun', text="No. Akun")
    tree.heading('nama_akun', text="Nama Akun")
    tree.heading('debet', text="Debet")
    tree.heading('kredit', text="Kredit")
    tree.heading('keterangan', text="Keterangan")
    tree.place(x=20, y=430)

    #Rekam Baris Jurnal Invoice
    rekam_invoice = tk.Button(frame_inv,text = "Rekam Invoice",bg="White",fg="black",font=["arial",12],width=15,command=rekam_baris)
    rekam_invoice.place(x=850,y=380)

    #generate Button
    generate_invoice = tk.Button(frame_inv,text = "Generate Invoice", bg="White",fg="black",font=["arial",14],width=40,command=inv_generate)
    generate_invoice.place(x=30, y=670)

    #new Invoice Button
    new_invoice = tk.Button(frame_inv, text = "New Invoice", bg="White",fg="black",font=["arial",14],width=40, command=inv_baru)
    new_invoice.place(x=30, y=720)

    #Print Invoice Button
    print_invoice= tk.Button(frame_inv,text = "Print Invoice", bg="White",fg="black",font=["arial",14],width=20, command = print_inv)
    print_invoice.place(x=600,y=700)

    window_invoice.mainloop