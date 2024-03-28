import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
import pandas as pd
import openpyxl as ox
from openpyxl import load_workbook
from tkinter import *
from tkcalendar import Calendar
import os
import locale
import datetime
import numpy as np
from num2words import num2words
import math
from pandastable import Table
import sys

# Receipting (payment)
def receipt ():
    window_receipt = tk.Tk()
    window_receipt.title("Receipting Entry")
    window_receipt.geometry("1080x780+400+10")
    window_receipt.configure(bg='#333333')
    frame_kwt=tk.Frame(window_receipt,bg='#333333')
    frame_kwt.pack(fill=BOTH,expand=True)
    #pilih tanggal formula
    def pick_date(event):
        global cal, date_window
        date_window = Toplevel()
        date_window.grab_set()
        date_window.title('Choose Date')
        date_window.geometry('250x220+590+370')
        cal = Calendar(date_window, selectmode="day", date_pattern="dd/mm/yyyy")
        cal.place(x=0, y=0)
        submit_btn = Button(date_window, text="submit", command=grab_date)
        submit_btn.place(x=80, y=190)
    def grab_date():
        date_entry.delete(0,END)
        date_entry.insert(0, cal.get_date())
        date_window.destroy()
    #customer formula
    def vlookup(value):
        filter_value = combobox_var.get()
        cust_kode['values'] = [value for value in cust_kode['values'] if isinstance(value, str) and value.startswith(filter_value)]
        result = custfile[custfile[custfile.columns[0]] == value][custfile.columns[1]].values[0]
        return result
    #invoice list formula
    def update_inv_list(event):
        selected_cust = cust_kode.get()
        if selected_cust:
            # Filter the invoice numbers based on the selected customer code
            inv_list = invfile[invfile.iloc[:, 1] == selected_cust].iloc[:, 0].tolist()
            inv_no.configure(values=inv_list)
    #bank formula
    def banklookup(value):
        filter_value1 = bankbox_var.get()
        bank_name['values'] = [value for value in bank_name['values'] if isinstance(value, str) and value.startswith(filter_value1)]
        result3 = bankfile[bankfile[bankfile.columns[1]] == value][bankfile.columns[0]].values[0]
        return result3
    #nilai Invoice
    payfile = pd.read_excel("Cust_file.xlsx",sheet_name="Data Pelunasan Inv")
    nilai_entry_var = tk.StringVar()
    def update_nilai_entry():
        selected_inv = inv_no.get()
        if selected_inv:
            result1= invfile[invfile[invfile.columns[0]] == selected_inv][invfile.columns[9]].values[0]
            result2 = payfile[payfile[payfile.columns[4]] == selected_inv].groupby(payfile.columns[4])[payfile.columns[6]].sum().values    
            if selected_inv in payfile[payfile.columns[4]].values and len(result2) > 0 and not np.isnan(result2[0]):
                result2 = int(result2[0])
                remaining_balance = int(result1 - result2)
                nilai_entry.delete(0, END)
                nilai_entry.insert(0, remaining_balance)
            else:
                nilai_entry.delete(0, END)
                nilai_entry.insert(0, int(result1))
    #akun ar       
    ar_entry_var = tk.StringVar()
    def update_ar_entry(event):
        selected_inv = inv_no.get()
        if selected_inv:
            result = invfile[invfile[invfile.columns[0]] == selected_inv][invfile.columns[10]].values[0]
            result2 = invfile[invfile[invfile.columns[0]] == selected_inv][invfile.columns[11]].values[0]
            ar_entry.delete(0, END)
            ar_entry.insert(0, result)
            nama_ar.delete(0,END)
            nama_ar.insert(0,result2)
    #terbilang formula
    def total_terbilang(value):
        try:
            payment_value = int(payment_entry.get())
            result4 = num2words(payment_value, lang="id")+" Rupiah"
            return result4
        except ValueError:
            return ''
    #Rekam Baris Formula
    receipt_list = []
    jurnal_list = []
    def rekam_baris():
        kode_kwt = "KWT"
        no_kwt = "KWT" + kwt_entry.get()
        kwt_no = kwt_entry.get()
        date = date_entry.get()
        cust_code = cust_kode.get()
        cust_name = custname_entry.get()
        no_inv = inv_no.get()
        sisa_inv = int(nilai_entry.get())
        value_pay = int(bayar_entry.get())
        value_inv = value_pay + int(pph23_entry.get())
        keterangan = keterangan_entry.get()
        transaksi = keterangan + cust_name + "atas" + no_inv
        terbilang = terbilang_entry.get()
        ar_akun = ar_entry.get()
        nama_piutang = nama_ar.get()
        pph23 = int(pph23_entry.get())
        #akun pph 23 prepaid
        if pph23 > 0 :
            akun_pph23 = "118-0004"
            date1 = date
            kode_kwt1=kode_kwt
            kwt_no1 = kwt_no
            transaksi1= transaksi
            keterangan1=keterangan
        else :
            akun_pph23 =""
            date1 = ""
            kode_kwt1=""
            kwt_no1 = ""
            transaksi1= ""
            keterangan1=""
        #nama akun pph23 prepaid
        if pph23 > 0:
            nama_pph23 = "PPh Pasal 23 Advance"
        else :
            nama_pph23 = ""            
        receipt_item = [no_kwt, date, cust_code, cust_name,no_inv ,sisa_inv , value_pay, pph23,keterangan, terbilang]
        jurnal_item_debet = [date1,kode_kwt1, kwt_no1,transaksi1,akun_pph23,nama_pph23,pph23,0,keterangan1]
        jurnal_item_kredit = [date,kode_kwt, kwt_no,transaksi,ar_akun,nama_piutang,0,value_inv,keterangan]
        tree.insert('',index= 'end', values=receipt_item)
        if pph23 >0 :
            jurnaltree.insert('',index= 'end', values=jurnal_item_kredit)
            jurnaltree.insert('',index= 'end', values=jurnal_item_debet)
        else :
            jurnaltree.insert('',index= 'end', values=jurnal_item_kredit)
        inv_no.delete(0,END)
        nilai_entry.delete(0, END)
        bayar_entry.delete(0,END)
        pph23_entry.delete(0,END)
        keterangan_entry.delete(0,END)
        ar_entry.delete(0,END)
        nama_ar.delete(0,END)
        receipt_list.append(receipt_item)
        jurnal_list.append(jurnal_item_kredit)
        jurnal_list.append(jurnal_item_debet)
    #Hapus Baris di treeview
    def hapus_baris():
        x = tree.selection()
        for record in x:
            tree.delete(record)
    #Receipt form baru setelah generate
    def new_kwt ():
        kwt_entry.delete(0,END)
        date_entry.delete(0,END)
        cust_kode.delete(0,END)
        custname_entry.delete(0,END)
        keterangan_entry.delete(0,END)
        akun_bank.delete(0, END)
        bank_name.delete(0,END)
        payment_entry.delete(0,END)
        terbilang_entry.delete(0,END)
        tree.delete(*tree.get_children())
        jurnaltree.delete(*jurnaltree.get_children())
    #Receipt form baru tanpa generate
    def kwt_baru ():
        confirmation1 = messagebox.askyesno("confirmation","apakah data sudah di generate?")
        if confirmation1 :
            kwt_entry.delete(0,END)
            date_entry.delete(0,END)
            cust_kode.delete(0,END)
            custname_entry.elete(0,END)
            keterangan_entry.delete(0,END)
            akun_bank.delete(0, END)
            bank_name.delete(0,END)
            payment_entry.delete(0,END)
            tree.delete(*tree.get_children())
            terbilang_entry.delete(0,END)
            jurnaltree.delete(*jurnaltree.get_children())
        else:
            confirmation2 = messagebox.askyesno("confirmation","apakan anda yakin akan membuat receipt baru?")
            if confirmation2:
                kwt_entry.delete(0,END)
                date_entry.delete(0,END)
                cust_kode.delete(0,END)
                custname_entry.delete(0,END)
                keterangan_entry.delete(0,END)
                akun_bank.delete(0, END)
                bank_name.delete(0,END)
                payment_entry.delete(0,END)
                terbilang_entry.delete(0,END)
                tree.delete(*tree.get_children())
                jurnaltree.delete(*jurnaltree.get_children())
            else :
                return
    #Generate Data ke database
    filepay2= ox.load_workbook('Cust_file.xlsx')
    ws2 = filepay2['Data Pelunasan Inv']
    def generate_receipt():
        #date = date_entry.get()
        #year = date[-4:]
        #file= f'database_{year}.xlsx'
        #if not os.path.exists(file):    
        #    workbook = ox.Workbook()
        #    ws = workbook.active
        #    heading = ["Tanggal,Kode Transaksi,No Transaksi,Transaksi,No Akun,Nama Akun,Debet,Kredit,Keterangan"]
        #    ws.append(heading)
        #    workbook.save(file)
        file = "database.xlsx"
        workbook = ox.load_workbook(file)
        ws = workbook.active
        no_kwt = kwt_entry.get()
        date = date_entry.get()
        cust_code = cust_kode.get()
        cust_name = custname_entry.get()
        transaksi = "Pembayaran "+ cust_name
        category = "oprt pay"
        bank_akun = akun_bank.get()
        nama_bank = bank_name.get()
        total_bayar = int(payment_entry.get())
        for receipt_item in tree.get_children():
            values = tree.item(receipt_item)["values"]
            ws2.append(values)
        ws.append([date,"KWT",no_kwt,transaksi,bank_akun,nama_bank,total_bayar,0,transaksi,category])
        for jurnal_item in jurnaltree.get_children():
            values1 = jurnaltree.item(jurnal_item)["values"]
            ws.append(values1)
        workbook.save(file)
        filepay2.save('Cust_file.xlsx')
        new_kwt()
    #Print Kwitansi
    template_kwitansi = ox.load_workbook('formkwitansi.xlsx')
    sheet = template_kwitansi['Form Kwitansi']
    def print_kwt():
        kwt_no = "KWT" + kwt_entry.get()
        name = custname_entry.get()
        sheet['j6'] = kwt_no
        sheet['i15'] = date_entry.get()
        sheet['d9'] = int(payment_entry.get())
        sheet['d7'] = name
        sheet['d11'] =terbilang_entry.get()
        sheet['d13'] ="Pembayaran "+ name
        new_file = kwt_no + name + " " + datetime.datetime.now().strftime("%Y-%m-%d")+".xlsx"
        template_kwitansi.save(new_file)
        os.startfile(new_file)
    
    #Design Kwitansi
    receipt_label = tk.Label(frame_kwt,text="RECEIPT",bg='#333333',fg='#FFFFFF',font=["arial",20],justify="center")
    receipt_label.pack(side=TOP)
    #Nomor kwitansi
    no_kwt = tk.Label(frame_kwt,text="Nomor Kwitansi",bg='#333333',fg='#FFFFFF',font=["arial",14],justify="left")
    no_kwt.place(x=20,y=70)
    kwt_kode = tk.Label(frame_kwt, text = "KWT", bg="white", fg="black",font=["arial",12])
    kwt_kode.place(x=180,y=70)
    kwt_entry = tk.Entry(frame_kwt, bg="white",fg="black",font=["arial",12],width=20)
    kwt_entry.place(x=230,y=70)
    #tanggal
    date_label = tk.Label(frame_kwt, text="Tanggal",bg ='#333333', fg='white',font=["arial",14])
    date_label.place(x=20, y=100)
    date_entry = tk.Entry(frame_kwt, highlightthickness=0,bg="white",fg="#6b6a69",font=("arial",14))
    date_entry.place(x=180, y=100)
    date_entry.insert(0, "dd/mm/yyyy")
    date_entry.bind("<1>",pick_date)
    #Customer
    custfile = pd.read_excel("Cust_file.xlsx", sheet_name="Data Base Cust")
    options = custfile[custfile.columns[0]].tolist()
    combobox_var = tk.StringVar()
    text_var = tk.StringVar()
    cust_label = tk.Label(frame_kwt,text="Customer",bg ='#333333', fg='white',font=["arial",14])
    cust_label.place(x=20, y=130)
    cust_kode = ttk.Combobox(frame_kwt,values=options,width=10,textvariable=combobox_var)
    cust_kode.place(x=30, y=160)
    cust_kode.bind("<<ComboboxSelected>>", lambda event:custname_entry.delete(0, END) or  custname_entry.insert(0, vlookup(cust_kode.get())))
    custname_entry = tk.Entry(frame_kwt, bg="white",fg="black",font=["arial",12],width=40)
    custname_entry.place(x=30,y=190)
    custname_entry.bind("<FocusOut>", update_inv_list)
    #Data Invoice & Total Payment
    invfile = pd.read_excel("Cust_file.xlsx", sheet_name="Data_Base_Invoice")
    inv_var = tk.StringVar()
    payment_label = tk.Label(frame_kwt, text="Total Payment",bg ='#333333', fg='white',font=["arial",12])
    payment_label.place(x=20, y=220)
    payment_entry = tk.Entry(frame_kwt,bg="white",fg="black",font=["arial",12])
    payment_entry.place(x=180,y=220)
    payment_entry.bind("<FocusOut>", lambda event: terbilang_entry.delete(0, END) or terbilang_entry.insert(0, total_terbilang(payment_entry.get())))
        #data Bank
    bankfile = pd.read_excel("lap_keu.xlsx",sheet_name="COA",usecols='A:B', skiprows=7, nrows=13)
    option_bank = bankfile[bankfile.columns[1]].tolist()
    bankbox_var = tk.StringVar()
    bank_label = tk.Label(frame_kwt, text = "Bank",bg ='#333333', fg='white',font=["arial",12])
    bank_label.place(x=420, y=220)
    bank_name = ttk.Combobox(frame_kwt,values=option_bank ,width=20, textvariable=bankbox_var)
    bank_name.place(x=480,y=220)
    bank_name.bind("<<ComboboxSelected>>",lambda event:akun_bank.delete(0, END) or  akun_bank.insert(0, banklookup(bank_name.get())))
    akun_bank = tk.Entry(frame_kwt,bg="white",fg="black",font=["arial",12])
        #data Invoice & Kwitansi
    invoice_label = tk.Label(frame_kwt,text="Invoice",bg ='#333333', fg='white',font=["arial",12])
    invoice_label.place(x=20, y=250)
    nilai_label = tk.Label(frame_kwt,text="Nilai Invoice",bg ='#333333', fg='white',font=["arial",12])
    nilai_label.place(x=180, y=250)
    inv_no = ttk.Combobox(frame_kwt, width=10, textvariable=inv_var)
    inv_no.place(x=20, y=280)
    inv_no_values = invfile.iloc[:,0].tolist()
    inv_no.configure(values=inv_no_values)
    inv_no.bind("<<ComboboxSelected>>",lambda event:(update_nilai_entry()))
    inv_no.bind("<FocusOut>", lambda event:(update_ar_entry(event)))
    ar_entry = tk.Entry(frame_kwt,bg="white",fg="black",font=["arial",12])
    nama_ar = tk.Entry(frame_kwt,bg="white",fg="black",font=["arial",12])
    nilai_entry = tk.Entry(frame_kwt,bg="white",fg="black",font=["arial",12])
    nilai_entry.place(x=180, y=280)
    nilai_bayar = tk.Label(frame_kwt,text="Nilai Bayar",bg ='#333333', fg='white',font=["arial",12])
    nilai_bayar.place(x=420, y=250)
    bayar_entry = tk.Entry(frame_kwt,bg="white",fg="black",font=["arial",12])
    bayar_entry.place(x=420, y=280)
    pph23_label = tk.Label(frame_kwt,text="PPH 23",bg ='#333333', fg='white',font=["arial",12])
    pph23_label.place(x=650, y=250)
    pph23_entry = tk.Entry(frame_kwt,bg="white",fg="black",font=["arial",12])
    pph23_entry.place(x=650, y=280)
    #keterangan
    keterangan_bayar = tk.Label(frame_kwt,text="keterangan",bg ='#333333', fg='white',font=["arial",12])
    keterangan_bayar.place(x=20, y=310)
    keterangan_entry = tk.Entry(frame_kwt,bg="white",fg="black",font=["arial",12],width=80)
    keterangan_entry.place(x=20, y=340)

    #Display payment table
    columns = ('no_kwt', 'tanggal', 'kode_cust', 'nama_cust','no_inv','nilai_inv', 'nilai_bayar','pph_23' ,'keterangan', 'terbilang')
    tree = ttk.Treeview(frame_kwt, columns=columns, show="headings")
    tree.heading('no_kwt', text="No. Kwitansi")
    tree.heading('tanggal', text="Tanggal Pembayaran")
    tree.heading('kode_cust', text="Kode Customer")
    tree.heading('nama_cust', text="Nama Customer")
    tree.heading('no_inv', text="No. Invoice")
    tree.heading('nilai_inv', text="Nilai sisa Invoice")
    tree.heading('nilai_bayar', text="Nominal Bayar")
    tree.heading('pph_23', text="PPH 23")
    tree.heading('keterangan', text="Keterangan")
    tree.heading('terbilang', text="Terbilang")
    tree.place(x=20, y=450,width=1000)
    treeXScroll = ttk.Scrollbar(tree, orient=HORIZONTAL, command=tree.xview)
    tree.configure(xscrollcommand=treeXScroll.set)
    treeXScroll.place(x=10, y=200,width=1000)

    #terbilang
    terbilang_label = tk.Label(frame_kwt,text="Terbilang",bg ='#333333', fg='white',font=["arial",12])
    terbilang_label.place(x=20, y=370)
    terbilang_entry = tk.Entry(frame_kwt,bg="white",fg="black",font=["arial",12],width=80)
    terbilang_entry.place(x=20, y=400)

    #Button
    rekam_button = tk.Button(frame_kwt,text= "Rekam Payment",bg="White",fg="black",font=["arial",12],width=25,command=rekam_baris)
    rekam_button.place(x=800 ,y=345)
    hapus_terpilih = tk.Button(frame_kwt,text= "Hapus Pilihan",bg="White",fg="black",font=["arial",12],width=25,command= hapus_baris)
    hapus_terpilih.place(x=800 ,y=385)
    generate_button = tk.Button(frame_kwt,text="Generate Payment",bg="White",fg="black",font=["arial",12],width=45,command=generate_receipt)
    generate_button.place(x=20,y=690)
    new_receipt = tk.Button(frame_kwt,text="Kwitansi Baru",bg="White",fg="black",font=["arial",12],width=45,command=kwt_baru)
    new_receipt.place(x=20,y=730)
    print_button = tk.Button(frame_kwt,text="Print Payment",bg="White",fg="black",font=["arial",12],width=45,command=print_kwt)
    print_button.place(x=480,y=700)

    #jurnal tree
    columns_jurnal = ('tanggal','kode_transaksi','no_kwt','transaksi', 'no_akun','nama_akun','debet', 'kredit','keterangan')
    jurnaltree = ttk.Treeview(frame_kwt, columns=columns_jurnal, show="headings")
    jurnaltree.heading('tanggal', text="Tanggal Pembayaran")
    jurnaltree.heading('kode_transaksi', text="Kode Transaksi")
    jurnaltree.heading('no_kwt', text="No. Kwitansi")
    jurnaltree.heading('transaksi', text="Transaksi")
    jurnaltree.heading('no_akun', text="No. Akun")
    jurnaltree.heading('nama_akun', text="Nama Akun")
    jurnaltree.heading('debet', text="Debet")
    jurnaltree.heading('kredit', text="Kredit")
    jurnaltree.heading('keterangan', text="Keterangan")
    #jurnaltree.place(x=20,y=690,width=1500)
    
    window_receipt.mainloop